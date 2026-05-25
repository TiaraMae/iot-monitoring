import os
import io
import sys
import json
import time
import math
import random
import threading
from datetime import datetime, timedelta, timezone

import psycopg2
import psycopg2.errors
import bcrypt
import paho.mqtt.client as mqtt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import numpy as np
import requests
from collections import defaultdict
from dotenv import load_dotenv

from flask import Flask, request, jsonify, render_template, redirect, url_for, flash, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user

app = Flask(__name__)
load_dotenv()

app.secret_key = os.getenv('FLASK_SECRET_KEY')
if not app.secret_key:
    raise RuntimeError("FLASK_SECRET_KEY environment variable is required")

# --- CONFIGURATION ---
MQTT_HOST = os.getenv("MQTT_HOST", "d57bf82836a7485d9b67b270c681fe6e.s1.eu.hivemq.cloud")
MQTT_PORT = int(os.getenv("MQTT_PORT", "8883"))
MQTT_USER = os.getenv("MQTT_USER", "esp32user")
MQTT_PASS = os.getenv("MQTT_PASS")
if not MQTT_PASS:
    raise RuntimeError("MQTT_PASS environment variable is required")

UNPAIRED_CACHE = {}
DEDUPE_CACHE = {}
EVENT_DEDUPE_CACHE = {}
CALIBRATION_TRACKER = {}  # appliance_id -> {start_tcoil, start_time}
# --- FAULT ALERT TRACKERS ---
FAULT_ALERT_TRACKER = {}      # appliance_id -> {fault_type: {cycle_count, last_trigger, active}}
FAULT_ALERT_COOLDOWN = {}     # (appliance_id, fault_type) -> last_alert_timestamp
DRYER_CYCLE_STATS = {}        # appliance_id -> {current_cycle: {start_time, motor_readings[], spike_peaks[], min_current}}
HVAC_CYCLE_TRACKER = {}       # appliance_id -> {state, start_time, best_reading, peak_current, maintain_start}

CLIENT_ID = f"FlaskBackend_{random.randint(10000, 99999)}"

# --- MQTT Setup ---
mqtt_client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2, client_id=CLIENT_ID, protocol=mqtt.MQTTv5)
mqtt_client.username_pw_set(MQTT_USER, MQTT_PASS)
mqtt_client.tls_set()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

DB_PASSWORD = os.getenv("DB_PASSWORD")
if not DB_PASSWORD:
    raise RuntimeError("DB_PASSWORD environment variable is required")

from psycopg2 import pool
try:
    db_pool = psycopg2.pool.SimpleConnectionPool(
        1, 10,
        host=os.getenv("DB_HOST", "localhost"),
        port=int(os.getenv("DB_PORT", "5432")),
        dbname=os.getenv("DB_NAME", "iot_db"),
        user=os.getenv("DB_USER", "postgres"),
        password=DB_PASSWORD
    )
except Exception as e:
    print(f"DB Pool Error: {e}")
    db_pool = None

def get_conn():
    try:
        if db_pool:
            return db_pool.getconn()
    except Exception as e:
        print(f"DB Get Connection Error: {e}")
    return None

def release_conn(conn):
    if db_pool and conn:
        db_pool.putconn(conn)


# --- STARTUP MIGRATIONS ---
def _run_startup_migrations():
    """Run lightweight idempotent migrations on startup."""
    conn = get_conn()
    if not conn:
        return
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'appliances' AND column_name = 'voltage'
        """)
        if not cur.fetchone():
            cur.execute("ALTER TABLE appliances ADD COLUMN voltage REAL DEFAULT 220.0")
            conn.commit()
            print("Migration applied: added 'voltage' column to appliances")
        # Add severity column to alerts table (idempotent)
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'alerts' AND column_name = 'severity'
        """)
        if not cur.fetchone():
            cur.execute("ALTER TABLE alerts ADD COLUMN severity VARCHAR(20) DEFAULT 'warning'")
            conn.commit()
            print("Migration applied: added 'severity' column to alerts")
    except Exception as e:
        print(f"Startup migration error: {e}")
    finally:
        cur.close()
        release_conn(conn)

_run_startup_migrations()


class User(UserMixin):
    def __init__(self, id, email, name):
        self.id = id
        self.email = email
        self.name = name

@login_manager.user_loader
def load_user(user_id):
    conn = get_conn()
    if not conn: return None
    cur = conn.cursor()
    cur.execute("SELECT id, email, name FROM users WHERE id = %s", (user_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    return User(row[0], row[1], row[2]) if row else None

# --- CALIBRATION HELPERS ---
def get_appliance_voltage(appliance_id):
    """Fetch appliance voltage (default 220.0 V for Indonesia)."""
    conn = get_conn()
    if not conn:
        return 220.0
    cur = conn.cursor()
    try:
        cur.execute("SELECT voltage FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        return float(row[0]) if row and row[0] is not None else 220.0
    except Exception:
        return 220.0
    finally:
        cur.close()
        release_conn(conn)


def _compute_daily_energy(readings, voltage):
    """Compute total energy (kWh) for a list of (time, current) readings.
    Cycles are split by >120s gaps or current dropping below 0.25A.
    """
    energy_ws = 0.0
    in_cycle = False
    cycle_readings = []
    for i, r in enumerate(readings):
        time_val, icompressor = r
        icompressor = float(icompressor) if icompressor is not None else 0.0
        if in_cycle and i > 0:
            gap = (time_val - readings[i-1][0]).total_seconds()
            if gap > 120:
                for j in range(1, len(cycle_readings)):
                    dt = (cycle_readings[j][0] - cycle_readings[j-1][0]).total_seconds()
                    energy_ws += cycle_readings[j-1][1] * voltage * dt
                in_cycle = False
                cycle_readings = []
        if icompressor > 0.25 and not in_cycle:
            in_cycle = True
            cycle_readings = [(time_val, icompressor)]
        elif in_cycle:
            if cycle_readings and cycle_readings[-1][0] != time_val:
                cycle_readings.append((time_val, icompressor))
        if icompressor < 0.25 and in_cycle:
            for j in range(1, len(cycle_readings)):
                dt = (cycle_readings[j][0] - cycle_readings[j-1][0]).total_seconds()
                energy_ws += cycle_readings[j-1][1] * voltage * dt
            in_cycle = False
            cycle_readings = []
    if in_cycle and cycle_readings:
        for j in range(1, len(cycle_readings)):
            dt = (cycle_readings[j][0] - cycle_readings[j-1][0]).total_seconds()
            energy_ws += cycle_readings[j-1][1] * voltage * dt
    return round(energy_ws / 3_600_000, 4)


def _compute_energy_kwh(readings, voltage):
    """Compute total energy (kWh) from (time, current) readings.
    Integrates current * voltage * dt over all valid consecutive points.
    A gap > 120s breaks the integration (appliance was off).
    Works for both HVAC and Dryer.
    """
    energy_ws = 0.0
    for i in range(1, len(readings)):
        prev_time, prev_current = readings[i-1]
        curr_time, curr_current = readings[i]
        gap = (curr_time - prev_time).total_seconds()
        if gap <= 120 and prev_current > 0.25:
            dt = gap
            energy_ws += prev_current * voltage * dt
    return round(energy_ws / 3_600_000, 4)


def get_appliance_calibration(appliance_id):
    conn = get_conn()
    default_cal = {
        'type': 'HVAC',
        't1_m': 1.0, 't1_c': 0.0,
        'h1_m': 1.0, 'h1_c': 0.0,
        't2_m': 1.0, 't2_c': 0.0,
        'h2_m': 1.0, 'h2_c': 0.0,
        'tcoil_c': 0.0, 'tcoil_m': 1.0,
    }
    if not conn: return default_cal
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT type, treturn_slope, treturn_intercept,
                   rhreturn_slope, rhreturn_intercept,
                   tsupply_slope, tsupply_intercept,
                   rhsupply_slope, rhsupply_intercept,
                   tcoil_offset, tcoil_slope
            FROM appliances WHERE id = %s
        """, (appliance_id,))
        row = cur.fetchone()
        cur.close()
        release_conn(conn)
        if not row: return default_cal
        return {
            'type': row[0],
            't1_m': row[1] if row[1] is not None else 1.0,
            't1_c': row[2] if row[2] is not None else 0.0,
            'h1_m': row[3] if row[3] is not None else 1.0,
            'h1_c': row[4] if row[4] is not None else 0.0,
            't2_m': row[5] if row[5] is not None else 1.0,
            't2_c': row[6] if row[6] is not None else 0.0,
            'h2_m': row[7] if row[7] is not None else 1.0,
            'h2_c': row[8] if row[8] is not None else 0.0,
            'tcoil_c': row[9] if row[9] is not None else 0.0,
            'tcoil_m': row[10] if row[10] is not None else 1.0,
        }
    except Exception as e:
        print(f"Error getting calibration: {e}")
        if conn: release_conn(conn)
        return default_cal

def apply_calibration(raw_val, m, c):
    if raw_val is None: return 0.0
    return (float(raw_val) * float(m)) + float(c)

# --- DISCORD ALERT HELPERS ---
def get_user_webhook(appliance_id):
    """Fetch the Discord webhook URL for the user who owns this appliance."""
    conn = get_conn()
    if not conn:
        return None
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT u.discord_webhook_url
            FROM users u
            JOIN appliances a ON a.user_id = u.id
            WHERE a.id = %s
        """, (appliance_id,))
        row = cur.fetchone()
        return row[0] if row and row[0] else None
    except Exception as e:
        print(f"Error fetching webhook: {e}")
        return None
    finally:
        cur.close()
        release_conn(conn)

def get_appliance_name(appliance_id):
    """Quick name lookup for an appliance."""
    conn = get_conn()
    if not conn:
        return None
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        return row[0] if row else None
    except Exception as e:
        print(f"Error fetching appliance name: {e}")
        return None
    finally:
        cur.close()
        release_conn(conn)

# --- Discord fault alert templates: maintenance-ticket style ---
FAULT_DISCORD_MAP = {
    'fault_dryer_incomplete_drying': {
        'title': '🔵 Clothes Not Fully Dried',
        'description': 'Dryer cycle completed but clothes retain excessive moisture.',
        'cause': 'Overloading, worn heating element, or short cycle.',
        'action': 'Reduce load size and run another cycle.',
    },
    'fault_dryer_roller_wear': {
        'title': '🟠 Barrel Roller Worn Out',
        'description': 'Motor is drawing more current than normal to maintain drum rotation.',
        'cause': 'Support rollers under the drum are worn, increasing mechanical friction.',
        'action': 'Inspect and replace drum support rollers.',
    },
    'fault_dryer_belt_snapped': {
        'title': '🔴 Belt Snapped',
        'description': 'Drive belt connecting motor to drum has broken or slipped off.',
        'cause': 'Age, overloading, or misalignment.',
        'action': 'Replace drive belt immediately.',
    },
    'fault_dryer_lint_blockage': {
        'title': '🔴 Lint Blockage Detected',
        'description': 'Lint accumulation is restricting exhaust airflow.',
        'cause': 'Failure to clean lint filter or exhaust duct; exterior vent obstruction.',
        'action': 'Clean lint filter and inspect exhaust duct.',
    },
    'fault_hvac_dirty_filter': {
        'title': '🟠 Dirty Indoor Filter',
        'description': 'Air filter is clogged, restricting airflow across the evaporator coil.',
        'cause': 'Neglected filter replacement; high dust environments.',
        'action': 'Replace or clean the indoor air filter.',
    },
    'fault_hvac_compressor_degradation': {
        'title': '🟠 Compressor Performance Degradation',
        'description': 'Compressor is drawing more current than normal but cooling output is below baseline.',
        'cause': 'Worn compressor valves, aging compressor, or gradual refrigerant loss reducing efficiency.',
        'action': 'Schedule HVAC technician to inspect compressor amp draw and refrigerant charge.',
    },
    'fault_hvac_low_refrigerant': {
        'title': '🔴 Low Refrigerant',
        'description': 'Refrigerant charge is below specification.',
        'cause': 'Micro-leaks in coil or lines; improper initial charge; Schrader valve leaks.',
        'action': 'Contact HVAC technician to check for leaks and recharge.',
    },
    'fault_hvac_compressor_fault': {
        'title': '🔴 Compressor Electrical Fault',
        'description': 'Compressor is drawing excessive current.',
        'cause': 'Failing compressor bearings, refrigerant overcharge, condenser blockage, or starter relay failure.',
        'action': 'Contact HVAC technician for compressor inspection.',
    },
}


def send_discord_alert(appliance_id, alert_type, message, value=None, threshold=None, severity='warning'):
    """Fire-and-forget Discord webhook alert. Non-blocking.
    Only fault alerts get the maintenance-ticket embed format."""
    try:
        webhook_url = get_user_webhook(appliance_id)
        if not webhook_url:
            return

        app_name = get_appliance_name(appliance_id)

        fault_meta = FAULT_DISCORD_MAP.get(alert_type)
        if fault_meta:
            # Severity-based color: critical = red, warning = amber/blue per fault type
            if severity == 'critical':
                embed_color = 0xEF4444  # red
                severity_prefix = "🚨 CRITICAL: "
            else:
                severity_prefix = "⚠️ WARNING: "
                warning_colors = {
                    'fault_dryer_incomplete_drying': 0x3B82F6,  # blue
                    'fault_dryer_roller_wear': 0xF59E0B,        # amber
                    'fault_hvac_dirty_filter': 0xF59E0B,        # amber
                    'fault_hvac_compressor_degradation': 0xF59E0B,  # amber
                }
                embed_color = warning_colors.get(alert_type, 0xF59E0B)
            embed = {
                "title": f"{severity_prefix}{fault_meta['title']}",
                "description": (
                    f"{message}\n\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"📍 **Appliance:** {app_name or f'ID {appliance_id}'}\n"
                    f"🔍 **Cause:** {fault_meta['cause']}\n"
                    f"🔧 **Recommended Action:** {fault_meta['action']}"
                ),
                "color": embed_color,
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "footer": {"text": "IoT Monitoring & Predictive Maintenance"}
            }
        else:
            # Generic fallback for any non-fault alert that still uses Discord
            embed = {
                "title": f"🚨 {alert_type.replace('_', ' ').title()}",
                "description": message,
                "color": 0x64748B,
                "fields": [
                    {"name": "Appliance", "value": app_name or f"ID {appliance_id}", "inline": True},
                    {"name": "Value", "value": str(value) if value is not None else "N/A", "inline": True},
                    {"name": "Threshold", "value": str(threshold) if threshold is not None else "N/A", "inline": True},
                ],
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "footer": {"text": "IoT Monitoring & Predictive Maintenance"}
            }

        resp = requests.post(webhook_url, json={"embeds": [embed]}, timeout=5)
        if resp.status_code not in (200, 204):
            print(f"Discord webhook returned {resp.status_code}: {resp.text}")
    except Exception as e:
        print(f"Discord alert failed: {e}")

# --- DATA QUERY HELPERS ---
def latest_row_for_appliance(appliance_id):
    conn = get_conn()
    if not conn: return None, None
    cur = conn.cursor()
    cur.execute("SELECT type, created_at FROM appliances WHERE id = %s", (appliance_id,))
    type_row = cur.fetchone()
    if not type_row:
        cur.close()
        release_conn(conn)
        return None, None
    dev_type, created_at = type_row[0], type_row[1]
    if "Dryer" in dev_type:
        cur.execute("""
            SELECT dr.time, dr.texhaust, dr.rh_exhaust, dr.pressure, dr.imotor
            FROM dryer_readings dr
            JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND dr.time >= a.created_at
            ORDER BY dr.time DESC LIMIT 1
        """, (appliance_id,))
    else:
        cur.execute("""
            SELECT sr.time, sr.treturn, sr.rhreturn, sr.tsupply, sr.rhsupply, sr.tcoil, sr.icompressor
            FROM hvac_readings sr
            JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND sr.time >= a.created_at
            ORDER BY sr.time DESC LIMIT 1
        """, (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    return row, dev_type

def latest_n_rows_for_appliance(appliance_id, limit, start=None, end=None, filtered=True):
    conn = get_conn()
    if not conn: return [], "Unknown"
    cur = conn.cursor()
    cur.execute("SELECT type, created_at FROM appliances WHERE id = %s", (appliance_id,))
    type_row = cur.fetchone()
    if not type_row: return [], "Unknown"
    dev_type, created_at = type_row[0], type_row[1]
    current_filter = " AND dr.imotor >= 0.25" if filtered else ""
    current_filter_hvac = " AND sr.icompressor >= 0.25" if filtered else ""
    if start and end:
        if "Dryer" in dev_type:
            cur.execute(f"""
                SELECT dr.time, dr.texhaust, dr.rh_exhaust, dr.pressure, dr.imotor
                FROM dryer_readings dr
                JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND dr.time >= %s AND dr.time <= %s AND dr.time >= %s{current_filter}
                ORDER BY dr.time ASC LIMIT 5000
            """, (appliance_id, start, end, created_at))
        else:
            cur.execute(f"""
                SELECT sr.time, sr.treturn, sr.rhreturn, sr.tsupply, sr.rhsupply, sr.tcoil, sr.icompressor
                FROM hvac_readings sr
                JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND sr.time >= %s AND sr.time <= %s AND sr.time >= %s{current_filter_hvac}
                ORDER BY sr.time ASC LIMIT 5000
            """, (appliance_id, start, end, created_at))
    else:
        if "Dryer" in dev_type:
            cur.execute(f"""
                SELECT dr.time, dr.texhaust, dr.rh_exhaust, dr.pressure, dr.imotor
                FROM dryer_readings dr
                JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND dr.time >= %s{current_filter}
                ORDER BY dr.time DESC LIMIT %s
            """, (appliance_id, created_at, limit))
        else:
            cur.execute(f"""
                SELECT sr.time, sr.treturn, sr.rhreturn, sr.tsupply, sr.rhsupply, sr.tcoil, sr.icompressor
                FROM hvac_readings sr
                JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND sr.time >= %s{current_filter_hvac}
                ORDER BY sr.time DESC LIMIT %s
            """, (appliance_id, created_at, limit))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return (rows[::-1], dev_type) if not (start and end) else (rows, dev_type)

def get_latest_current_for_appliance(appliance_id):
    conn = get_conn()
    if not conn: return 0.0
    cur = conn.cursor()
    try:
        cur.execute("SELECT type, created_at FROM appliances WHERE id = %s", (appliance_id,))
        type_row = cur.fetchone()
        if not type_row: return 0.0
        dev_type, created_at = type_row[0], type_row[1]
        if "Dryer" in dev_type:
            cur.execute("""
                SELECT dr.imotor FROM dryer_readings dr
                JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND dr.time >= a.created_at
                ORDER BY dr.time DESC LIMIT 1
            """, (appliance_id,))
        else:
            cur.execute("""
                SELECT sr.icompressor FROM hvac_readings sr
                JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND sr.time >= a.created_at
                ORDER BY sr.time DESC LIMIT 1
            """, (appliance_id,))
        row = cur.fetchone()
        return float(row[0]) if row and row[0] is not None else 0.0
    except Exception:
        return 0.0
    finally:
        cur.close()
        release_conn(conn)

def get_appliances_for_user(user_id):
    conn = get_conn()
    if not conn: return []
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, type, brand, location, created_at, operational_status, sub_type, baseline_configured
        FROM appliances WHERE user_id = %s ORDER BY created_at
    """, (user_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    appliances = []
    for r in rows:
        app = {'id':r[0],'name':r[1],'type':r[2],'brand':r[3],'location':r[4],
               'created_at':r[5], 'status':r[6], 'sub_type':r[7], 'baseline_configured':r[8],
               'alert_status': get_appliance_alert_status(r[0])}
        appliances.append(app)
    return appliances

def get_appliance_alert_status(appliance_id):
    """Return the highest severity of active (unresolved) alerts for an appliance.
    Returns one of: 'normal', 'warning', 'critical'."""
    conn = get_conn()
    if not conn:
        return 'normal'
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT severity FROM alerts
            WHERE appliance_id = %s AND resolved_at IS NULL
            ORDER BY CASE severity WHEN 'critical' THEN 3 WHEN 'warning' THEN 2 ELSE 1 END DESC
            LIMIT 1
        """, (appliance_id,))
        row = cur.fetchone()
        if row and row[0]:
            return row[0]
        return 'normal'
    except Exception as e:
        print(f"get_appliance_alert_status error: {e}")
        return 'normal'
    finally:
        cur.close()
        release_conn(conn)

def unpaired_nodes():
    conn = get_conn()
    if not conn: return []
    cur = conn.cursor()
    cur.execute("""
        SELECT sn.id, sn.mac_address, sn.status
        FROM sensor_nodes sn
        WHERE sn.status = 'unpaired' AND sn.last_seen >= NOW() - INTERVAL '30 seconds'
        ORDER BY sn.id
    """)
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return [{'id':r[0], 'mac_address':r[1], 'status':r[2], 'readings':{}} for r in rows]

def get_all_nodes_for_user(user_id):
    conn = get_conn()
    if not conn: return []
    cur = conn.cursor()
    cur.execute("""
        SELECT sn.id, sn.mac_address, sn.status, a.name
        FROM sensor_nodes sn
        LEFT JOIN appliances a ON sn.appliance_id = a.id
        WHERE a.user_id = %s OR sn.status = 'unpaired'
        ORDER BY sn.id
    """, (user_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return [{'id':r[0], 'mac_address':r[1], 'status':r[2], 'appliance':r[3] or 'Unpaired'} for r in rows]

def send_node_command(mac, command_str):
    if not mqtt_client.is_connected():
        print(f"Cannot send command {command_str} to {mac}, MQTT disconnected.")
        return
    mqtt_client.publish(f"iot/nodes/{mac}/control", command_str)
    print(f"Backend -> Node {mac}: {command_str}")

# --- SPC BASELINE HELPERS ---
HVAC_METRICS = ['deltat', 'current']
DRYER_METRICS = ['texhaust', 'rhexhaust', 'current']

def get_spc_baselines(appliance_id):
    """Fetch manual SPC baselines from spc_manual_baselines table."""
    conn = get_conn()
    if not conn: return {}
    cur = conn.cursor()
    cur.execute("""
        SELECT metric_name, ucl, lcl, mean FROM spc_manual_baselines WHERE appliance_id = %s
    """, (appliance_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    result = {}
    for r in rows:
        result[r[0]] = {'ucl': float(r[1]), 'lcl': float(r[2]), 'mean': float(r[3])}
    return result

def save_spc_baselines(appliance_id, baselines):
    """baselines: dict of metric_name -> {ucl, lcl}"""
    conn = get_conn()
    if not conn: return False, "DB connection error"
    cur = conn.cursor()
    try:
        for metric, vals in baselines.items():
            ucl = float(vals['ucl'])
            lcl = float(vals['lcl'])
            if ucl <= lcl:
                return False, f"UCL must be greater than LCL for {metric}"
            mean = (ucl + lcl) / 2.0
            cur.execute("""
                INSERT INTO spc_manual_baselines (appliance_id, metric_name, ucl, lcl, mean, updated_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
                ON CONFLICT (appliance_id, metric_name) DO UPDATE SET
                    ucl = EXCLUDED.ucl,
                    lcl = EXCLUDED.lcl,
                    mean = EXCLUDED.mean,
                    updated_at = NOW()
            """, (appliance_id, metric, ucl, lcl, mean))
        cur.execute("UPDATE appliances SET baseline_configured = TRUE WHERE id = %s", (appliance_id,))
        conn.commit()
        return True, "Baseline saved successfully"
    except Exception as e:
        conn.rollback()
        return False, str(e)
    finally:
        cur.close()
        release_conn(conn)

def notify_node_baseline_set(appliance_id):
    conn = get_conn()
    if not conn: return
    cur = conn.cursor()
    cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s", (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if row and row[0]:
        send_node_command(row[0], "baseline:set")

# --- FAULT ALERT SYSTEM ---

def _insert_fault_alert(appliance_id, alert_type, message, value, threshold, severity, now, cur, conn):
    """Insert a fault alert with 10-minute cooldown per fault type.
    severity: 'warning' or 'critical'"""
    cooldown_key = (appliance_id, alert_type)
    if cooldown_key in FAULT_ALERT_COOLDOWN:
        if (now - FAULT_ALERT_COOLDOWN[cooldown_key]).total_seconds() < 600:
            return
    try:
        cur.execute("""
            INSERT INTO alerts (appliance_id, alert_type, message, value, threshold, severity, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (appliance_id, alert_type, message, value, threshold, severity, now))
        conn.commit()
        FAULT_ALERT_COOLDOWN[cooldown_key] = now
        send_discord_alert(appliance_id, alert_type, message, value, threshold, severity)
        tracker = FAULT_ALERT_TRACKER.setdefault(appliance_id, {})
        ft = tracker.setdefault(alert_type, {'cycle_count': 0, 'last_trigger': now, 'active': True})
        ft['last_trigger'] = now
        ft['active'] = True
    except Exception as e:
        print(f"Fault alert insert error: {e}")


def _check_dryer_faults(appliance_id, reading_data, baselines, now, cur, conn):
    """Real-time dryer fault detection with per-cycle median tracking."""
    current = reading_data.get('current', 0.0)
    texhaust = reading_data.get('texhaust', 0.0)
    rhexhaust = reading_data.get('rhexhaust', 0.0)
    actual_time = reading_data.get('_actual_time', now)

    stats = DRYER_CYCLE_STATS.setdefault(appliance_id, {})
    mean_current = baselines.get('current', {}).get('mean', 2.0)
    hard_threshold = mean_current * 1.15

    # Check for cycle end due to gap (>120s)
    if stats.get('in_cycle', False) and 'last_time' in stats:
        gap = (actual_time - stats['last_time']).total_seconds()
        if gap > 120:
            _finalize_dryer_cycle(appliance_id, baselines, now, cur, conn)
            stats = DRYER_CYCLE_STATS[appliance_id] = {}

    # Start new cycle
    if current > 0.25 and not stats.get('in_cycle', False):
        stats = DRYER_CYCLE_STATS[appliance_id] = {
            'in_cycle': True,
            'start_time': actual_time,
            'last_time': actual_time,
            'motor_readings': [current],
            'spike_peaks': [],
            'spike_state': 'IDLE',
            'spike_max': 0.0,
            'spike_valley': 0.0,
            'prev_current': current,
            'min_current': current,
            'max_temp': texhaust,
            'temp_history': [texhaust],
            'rh_history': [rhexhaust] if rhexhaust is not None else [],
            'consecutive_below_lcl': 0,
            'belt_snap_triggered': False,
        }
        return  # Skip rest of processing for first reading (matches historical dryer_analytics)

    # During active cycle
    if stats.get('in_cycle', False):
        stats['last_time'] = actual_time
        stats['max_temp'] = max(stats.get('max_temp', texhaust), texhaust)
        stats.setdefault('temp_history', []).append(texhaust)
        if rhexhaust is not None:
            stats.setdefault('rh_history', []).append(rhexhaust)
        stats['min_current'] = min(stats.get('min_current', 999.0), current)

        prev = stats.get('prev_current', 0.0)
        prominence = 0.40
        state = stats.get('spike_state', 'IDLE')
        spike_max = stats.get('spike_max', 0.0)
        spike_valley = stats.get('spike_valley', 0.0)

        # Spike state machine (matches historical dryer_analytics logic)
        if current > prev:
            if state == 'FALLING':
                # Confirm previous peak before starting new rise
                if spike_max > 0:
                    prom = spike_max - spike_valley
                    if prom >= prominence and spike_max > mean_current + 0.15:
                        stats.setdefault('spike_peaks', []).append(spike_max)
                spike_max = 0.0
                spike_valley = 0.0
            if state in ('IDLE', 'FALLING'):
                spike_valley = prev
            state = 'RISING'
            if current > spike_max:
                spike_max = current
        elif current < prev:
            if state == 'RISING' and (spike_max <= 0.1 or current < spike_max - 0.1):
                state = 'FALLING'

        stats['spike_state'] = state
        stats['spike_max'] = spike_max
        stats['spike_valley'] = spike_valley

        # Collect all readings for motor baseline median calculation
        stats.setdefault('motor_readings', []).append(current)

        stats['prev_current'] = current

        # --- Belt snap detection (real-time) ---
        # Only runs on running data (final_amps >= 0.25), so idle/pause gaps are excluded
        current_lcl = baselines.get('current', {}).get('lcl')
        current_ucl = baselines.get('current', {}).get('ucl')

        if current_lcl is not None:
            if current < current_lcl:
                stats['consecutive_below_lcl'] = stats.get('consecutive_below_lcl', 0) + 1
            else:
                stats['consecutive_below_lcl'] = 0

            if stats['consecutive_below_lcl'] >= 3 and not stats.get('belt_snap_triggered', False):
                _insert_fault_alert(
                    appliance_id, 'fault_dryer_belt_snapped',
                    f"Belt snapped - motor current dropped below LCL {current_lcl:.3f}A for 3 consecutive readings (last: {current:.3f}A)",
                    current, current_lcl, 'critical', now, cur, conn)
                stats['belt_snap_triggered'] = True

    # Cycle end by current drop
    if current < 0.15 and stats.get('in_cycle', False):
        _finalize_dryer_cycle(appliance_id, baselines, now, cur, conn)


def _compute_motor_baseline_median(motor_readings, filter_threshold=None):
    """Compute median of motor baseline readings, optionally filtering outliers above filter_threshold."""
    if not motor_readings:
        return 0.0
    readings = motor_readings
    if filter_threshold is not None:
        readings = [r for r in readings if r <= filter_threshold]
        if not readings:
            return 0.0
    sorted_r = sorted(readings)
    n = len(sorted_r)
    if n % 2 == 1:
        return sorted_r[n // 2]
    return (sorted_r[n // 2 - 1] + sorted_r[n // 2]) / 2.0


def _finalize_dryer_cycle(appliance_id, baselines, now, cur, conn):
    """Finalize a dryer cycle and evaluate end-of-cycle faults."""
    stats = DRYER_CYCLE_STATS.get(appliance_id, {})
    if not stats or not stats.get('in_cycle', False):
        return

    # Gate on baseline configured and alerts enabled
    cur.execute("SELECT baseline_configured, alert_enabled FROM appliances WHERE id = %s", (appliance_id,))
    app_info = cur.fetchone()
    if not app_info or not app_info[0] or not app_info[1]:
        DRYER_CYCLE_STATS[appliance_id] = {}
        return

    motor_readings = stats.get('motor_readings', [])
    rh_history = stats.get('rh_history', [])
    max_temp = stats.get('max_temp', 0.0)

    # Confirm any pending spike before finalizing
    mean_current = baselines.get('current', {}).get('mean', 2.0)
    prominence = 0.40
    spike_state = stats.get('spike_state', 'IDLE')
    spike_max = stats.get('spike_max', 0.0)
    spike_valley = stats.get('spike_valley', 0.0)
    if spike_state in ('RISING', 'FALLING') and spike_max > 0:
        prom = spike_max - spike_valley
        if prom >= prominence and spike_max > mean_current + 0.15:
            stats.setdefault('spike_peaks', []).append(spike_max)

    # Compute begin-of-cycle RH (first 6 readings or first 1 min equivalent)
    begin_rh_avg = 0.0
    if rh_history:
        first_rh = rh_history[:6] if len(rh_history) > 6 else rh_history
        begin_rh_avg = sum(first_rh) / len(first_rh)

    # Compute end-of-cycle RH (last 6 readings)
    end_rh_avg = 0.0
    if rh_history:
        last_rh = rh_history[-6:] if len(rh_history) > 6 else rh_history
        end_rh_avg = sum(last_rh) / len(last_rh)

    # Compute end-of-cycle temp (last 6 readings)
    temp_history = stats.get('temp_history', [])
    end_temp_avg = 0.0
    if temp_history:
        last_temp = temp_history[-6:] if len(temp_history) > 6 else temp_history
        end_temp_avg = sum(last_temp) / len(last_temp)

    # --- Lint Blockage ---
    rhexhaust_ucl = baselines.get('rhexhaust', {}).get('ucl')
    texhaust_ucl = baselines.get('texhaust', {}).get('ucl')
    if rhexhaust_ucl is not None and texhaust_ucl is not None:
        if end_rh_avg > rhexhaust_ucl and max_temp > texhaust_ucl:
            # Critical if burning risk (temp > 100C), otherwise warning
            if max_temp > 100.0:
                severity = 'critical'
                msg = f"🔥 BURNING RISK - Lint blockage detected! End RH {end_rh_avg:.1f}% > UCL {rhexhaust_ucl:.1f}% and exhaust temp {max_temp:.1f}C > UCL {texhaust_ucl:.1f}C (CRITICAL: temp > 100C)"
            else:
                severity = 'warning'
                msg = f"Lint blockage detected - end RH {end_rh_avg:.1f}% > UCL {rhexhaust_ucl:.1f}% and exhaust temp {max_temp:.1f}C > UCL {texhaust_ucl:.1f}C"
            _insert_fault_alert(
                appliance_id, 'fault_dryer_lint_blockage',
                msg, end_rh_avg, rhexhaust_ucl, severity, now, cur, conn)

    # --- Incomplete Drying ---
    # Check even if lint blockage fired (independent alert)
    if rhexhaust_ucl is not None and end_rh_avg > rhexhaust_ucl:
        if end_rh_avg > 90.0:
            severity = 'critical'
            msg = f"Severely incomplete drying - end RH {end_rh_avg:.1f}% exceeds 90% (not drying at all)"
        else:
            severity = 'warning'
            msg = f"Clothes not fully dried - end RH {end_rh_avg:.1f}% exceeds UCL {rhexhaust_ucl:.1f}%"
        _insert_fault_alert(
            appliance_id, 'fault_dryer_incomplete_drying',
            msg, end_rh_avg, rhexhaust_ucl, severity, now, cur, conn)

    # --- Belt Snap (end-of-cycle backup) ---
    current_lcl = baselines.get('current', {}).get('lcl')
    if current_lcl is not None and not stats.get('belt_snap_triggered', False):
        last_3 = motor_readings[-3:] if len(motor_readings) >= 3 else motor_readings
        if len(last_3) >= 3 and all(r < current_lcl for r in last_3):
            _insert_fault_alert(
                appliance_id, 'fault_dryer_belt_snapped',
                f"Belt snapped - last 3 motor readings ({last_3[-3]:.3f}A, {last_3[-2]:.3f}A, {last_3[-1]:.3f}A) all below LCL {current_lcl:.3f}A",
                last_3[-1], current_lcl, 'critical', now, cur, conn)

    # --- Roller Wear (end-of-cycle backup) ---
    current_ucl = baselines.get('current', {}).get('ucl')
    if current_ucl is not None and not stats.get('roller_wear_triggered', False):
        if motor_readings:
            filter_threshold = (sum(motor_readings) / len(motor_readings)) * 1.15
            median_current = _compute_motor_baseline_median(motor_readings, filter_threshold=filter_threshold)
            if median_current > current_ucl:
                _insert_fault_alert(
                    appliance_id, 'fault_dryer_roller_wear',
                    f"Barrel roller worn out - motor baseline median {median_current:.3f}A exceeded UCL {current_ucl:.3f}A during cycle",
                    median_current, current_ucl, 'warning', now, cur, conn)

    # Clear cycle stats
    DRYER_CYCLE_STATS[appliance_id] = {}


def _check_hvac_faults(appliance_id, reading_data, baselines, now, cur, conn, is_inverter):
    """HVAC fault detection with last-6-reading window evaluation (~1 min of data).
    Non-inverter: evaluates at 1 hour, then re-evaluates every additional hour while running.
    Inverter: evaluates at 5 minutes when Treturn > 26.5C.
    Compressor fault (current > UCL) triggers IMMEDIATELY on any single reading."""
    current = reading_data.get('current', 0.0)
    deltat = reading_data.get('deltat', 0.0)
    treturn = reading_data.get('treturn', 0.0)

    tracker = HVAC_CYCLE_TRACKER.setdefault(appliance_id, {
        'state': 'IDLE',
        'start_time': None,
        'reading_buffer': [],
        'last_evaluation_runtime': 0,
    })

    def _add_to_buffer():
        buf = tracker['reading_buffer']
        buf.append({'deltat': deltat, 'current': current, 'treturn': treturn})
        if len(buf) > 6:
            buf.pop(0)

    def _runtime_seconds():
        if tracker['start_time'] is None:
            return 0
        return (now - tracker['start_time']).total_seconds()

    def _evaluate_if_ready(min_runtime, force=False):
        """Evaluate if we've passed min_runtime since start or last evaluation."""
        runtime = _runtime_seconds()
        last_eval = tracker['last_evaluation_runtime']
        if not tracker['reading_buffer']:
            return
        if force or runtime >= min_runtime:
            if last_eval == 0 or (runtime - last_eval) >= min_runtime:
                _evaluate_hvac_window(
                    appliance_id, tracker['reading_buffer'], baselines,
                    runtime, now, cur, conn)
                tracker['last_evaluation_runtime'] = runtime

    def _reset_tracker():
        tracker['state'] = 'IDLE'
        tracker['start_time'] = None
        tracker['reading_buffer'] = []
        tracker['last_evaluation_runtime'] = 0

    # --- IMMEDIATE COMPRESSOR FAULT: current > UCL on ANY single reading ---
    current_ucl = baselines.get('current', {}).get('ucl')
    if current_ucl is not None and current > current_ucl:
        _insert_fault_alert(
            appliance_id, 'fault_hvac_compressor_fault',
            f"Outdoor problem - Current {current:.2f}A exceeded UCL {current_ucl:.2f}A",
            current, current_ucl, 'critical', now, cur, conn)

    if not is_inverter:
        # --- NON-INVERTER ---
        if current > 0.25:
            if tracker['state'] == 'IDLE':
                tracker['state'] = 'RUNNING'
                tracker['start_time'] = now
                tracker['reading_buffer'] = [{'deltat': deltat, 'current': current, 'treturn': treturn}]
                tracker['last_evaluation_runtime'] = 0
            elif tracker['state'] == 'RUNNING':
                _add_to_buffer()
                # Evaluate at 1 hour, then every additional hour
                _evaluate_if_ready(min_runtime=3600)
        else:
            if tracker['state'] == 'RUNNING':
                # Final evaluation before turn-off if never evaluated this cycle
                if tracker['last_evaluation_runtime'] == 0 and tracker['reading_buffer']:
                    _evaluate_hvac_window(
                        appliance_id, tracker['reading_buffer'], baselines,
                        _runtime_seconds(), now, cur, conn)
            _reset_tracker()
    else:
        # --- INVERTER ---
        is_high_effort = (current > 0.25 and treturn > 26.5)
        if is_high_effort:
            if tracker['state'] == 'IDLE':
                tracker['state'] = 'RUNNING'
                tracker['start_time'] = now
                tracker['reading_buffer'] = [{'deltat': deltat, 'current': current, 'treturn': treturn}]
                tracker['last_evaluation_runtime'] = 0
            elif tracker['state'] == 'RUNNING':
                _add_to_buffer()
                # Evaluate at 5 minutes
                _evaluate_if_ready(min_runtime=300)
        else:
            if tracker['state'] == 'RUNNING':
                # Final evaluation before drop-off if never evaluated this cycle
                if tracker['last_evaluation_runtime'] == 0 and tracker['reading_buffer']:
                    _evaluate_hvac_window(
                        appliance_id, tracker['reading_buffer'], baselines,
                        _runtime_seconds(), now, cur, conn)
            _reset_tracker()


def _evaluate_hvac_window(appliance_id, reading_buffer, baselines, runtime, now, cur, conn):
    """Evaluate HVAC faults using the average of the last 6 buffered readings.
    Fires at most ONE alert per cycle/window."""
    if not reading_buffer:
        return

    # Compute averages from last 6 readings
    deltat_avg = sum(r['deltat'] for r in reading_buffer) / len(reading_buffer)
    current_avg = sum(r['current'] for r in reading_buffer) / len(reading_buffer)
    treturn_avg = sum(r['treturn'] for r in reading_buffer) / len(reading_buffer)

    deltat_lcl = baselines.get('deltat', {}).get('lcl')
    current_lcl = baselines.get('current', {}).get('lcl')
    current_ucl = baselines.get('current', {}).get('ucl')
    current_mean = baselines.get('current', {}).get('mean')

    if deltat_lcl is None or current_lcl is None or current_ucl is None or current_mean is None:
        return  # Baselines not fully configured

    if deltat_avg >= deltat_lcl:
        # Delta-T is large enough — good cooling performance, no fault
        return

    # Step 1: Determine severity based on runtime + Treturn
    if runtime >= 600 and treturn_avg >= 27.0:
        severity = 'critical'
    else:
        severity = 'warning'

    # Step 2: Determine fault type from current matrix
    # Upper warning zone threshold = mean + 1σ (where σ = (UCL-mean)/3)
    current_warn = current_mean + (current_ucl - current_mean) / 3.0

    if current_avg < current_lcl:
        _insert_fault_alert(
            appliance_id, 'fault_hvac_low_refrigerant',
            f"Low refrigerant - Delta-T {deltat_avg:.1f}C below LCL {deltat_lcl:.1f}C with low current {current_avg:.2f}A",
            deltat_avg, deltat_lcl, severity, now, cur, conn)
    elif current_avg > current_ucl:
        _insert_fault_alert(
            appliance_id, 'fault_hvac_compressor_fault',
            f"Outdoor problem - Delta-T {deltat_avg:.1f}C below LCL {deltat_lcl:.1f}C with high current {current_avg:.2f}A",
            current_avg, current_ucl, severity, now, cur, conn)
    elif current_avg > current_warn:
        _insert_fault_alert(
            appliance_id, 'fault_hvac_compressor_degradation',
            f"Compressor degradation - Delta-T {deltat_avg:.1f}C below LCL {deltat_lcl:.1f}C with elevated current {current_avg:.2f}A (baseline mean {current_mean:.2f}A)",
            current_avg, current_mean, severity, now, cur, conn)
    else:
        _insert_fault_alert(
            appliance_id, 'fault_hvac_dirty_filter',
            f"Dirty indoor filter - Delta-T {deltat_avg:.1f}C below LCL {deltat_lcl:.1f}C with normal current {current_avg:.2f}A",
            deltat_avg, deltat_lcl, severity, now, cur, conn)


def check_fault_alerts(appliance_id, reading_data, dev_type, actual_time):
    """Check for pattern-based fault alerts. Gated behind baseline_configured."""
    baselines = get_spc_baselines(appliance_id)
    if not baselines:
        return
    conn = get_conn()
    if not conn:
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT baseline_configured, alert_enabled, sub_type FROM appliances WHERE id = %s", (appliance_id,))
        app_info = cur.fetchone()
        if not app_info or not app_info[0] or not app_info[1]:
            return
        baseline_configured, alert_enabled, sub_type = app_info[0], app_info[1], app_info[2]
        if not baseline_configured or not alert_enabled:
            return

        now = datetime.now(timezone.utc)
        reading_data['_actual_time'] = actual_time
        is_inverter = (sub_type == 'inverter')

        if "Dryer" in dev_type:
            _check_dryer_faults(appliance_id, reading_data, baselines, now, cur, conn)
        else:
            _check_hvac_faults(appliance_id, reading_data, baselines, now, cur, conn, is_inverter)
    except Exception as e:
        print(f"Fault alert check error: {e}")
    finally:
        if conn:
            cur.close()
            release_conn(conn)


# --- MQTT HANDLERS ---
def on_mqtt_connect(client, userdata, flags, rc, properties=None):
    if rc == 0:
        print(f"MQTT Connected Successfully! ID: {client._client_id}")
        client.subscribe("iot/nodes/+/events")
        client.subscribe("iot/nodes/+/telemetry")
    else:
        print(f"MQTT Connection Failed with Code {rc}")

def handle_node_events(mac, payload):
    try:
        data = json.loads(payload)
        event_type = data.get("event")
        event_hash = f"{mac}_{event_type}"
        now_time = time.time()
        if event_hash in EVENT_DEDUPE_CACHE:
            if now_time - EVENT_DEDUPE_CACHE[event_hash] < 5.0:
                print(f"Duplicate event {event_type} from {mac} ignored (in-memory dedup).")
                return
        EVENT_DEDUPE_CACHE[event_hash] = now_time

        conn = get_conn()
        if not conn: return
        cur = conn.cursor()
        cur.execute("SELECT appliance_id, status FROM sensor_nodes WHERE mac_address = %s", (mac,))
        node_row = cur.fetchone()

        if event_type == "event_request_config":
            if not node_row:
                cur.execute(
                    "INSERT INTO sensor_nodes (mac_address, status, created_at, last_seen) VALUES (%s, 'unpaired', NOW(), NOW())",
                    (mac,)
                )
                conn.commit()
                print(f"Auto-registered unknown node {mac}")
            else:
                cur.execute(
                    "UPDATE sensor_nodes SET last_seen = NOW() WHERE mac_address = %s",
                    (mac,)
                )
                conn.commit()
            if not node_row or not node_row[0] or (node_row and node_row[1] != 'paired'):
                send_node_command(mac, "settype:unpaired")
            else:
                appliance_id = node_row[0]
                cur.execute("SELECT type, operational_status, cf, deductor FROM appliances WHERE id = %s", (appliance_id,))
                row = cur.fetchone()
                if row:
                    app_type, app_status = row[0], row[1]
                    cf = row[2] if len(row) > 2 and row[2] is not None else (33.0 if "Dryer" in app_type else 11.0)
                    deductor = row[3] if len(row) > 3 and row[3] is not None else (0.111 if "Dryer" in app_type else 0.033)
                    if app_status == 'calibrating':
                        cur.execute("UPDATE appliances SET operational_status='calibration_needed' WHERE id=%s", (appliance_id,))
                        conn.commit()
                        app_status = 'calibration_needed'
                    if "Dryer" in app_type:
                        send_node_command(mac, "settype:dryer")
                        send_node_command(mac, f"setcf:{cf}")
                        send_node_command(mac, f"setdeductor:{deductor}")
                        send_node_command(mac, "restore:normal")
                    else:
                        send_node_command(mac, "settype:hvac")
                        send_node_command(mac, f"setcf:{cf}")
                        send_node_command(mac, f"setdeductor:{deductor}")
                        if app_status == 'normal':
                            send_node_command(mac, "restore:normal")
                        else:
                            send_node_command(mac, "restore:calibrationneeded")
            cur.close()
            release_conn(conn)
            return

        if event_type == "checkin":
            cur.execute("UPDATE sensor_nodes SET last_seen = NOW() WHERE mac_address = %s", (mac,))
            conn.commit()
            cur.close()
            release_conn(conn)
            return

        if not node_row or not node_row[0] or node_row[1] != 'paired':
            if event_type in ["event_button2_action_request"]:
                send_node_command(mac, "actiondenied:busy")
            else:
                send_node_command(mac, "baselinefailack")
            cur.close()
            release_conn(conn)
            return

        appliance_id = node_row[0]
        cur.execute("SELECT operational_status, type FROM appliances WHERE id = %s", (appliance_id,))
        stat_row = cur.fetchone()
        if not stat_row:
            cur.close()
            release_conn(conn)
            return
        status, app_type = stat_row[0], stat_row[1]

        if event_type == "event_button2_calibration_request":
            if "Dryer" in app_type:
                send_node_command(mac, "calibrationfailack")
                print(f"Node {mac} (Dryer) Calibration not supported.")
            else:
                if status == 'calibration_needed':
                    cur.execute("UPDATE appliances SET operational_status = 'calibrating', calibration_started_at = NOW() WHERE id = %s", (appliance_id,))
                    conn.commit()
                    cur.execute("""
                        SELECT sr.tcoil FROM hvac_readings sr
                        JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                        WHERE sn.appliance_id = %s ORDER BY sr.time DESC LIMIT 1
                    """, (appliance_id,))
                    tcoil_row = cur.fetchone()
                    start_tcoil = float(tcoil_row[0]) if tcoil_row and tcoil_row[0] is not None else 25.0
                    CALIBRATION_TRACKER[appliance_id] = {'start_tcoil': start_tcoil, 'start_time': datetime.now()}
                    send_node_command(mac, "startcalibration")
                    print(f"Node {mac} (HVAC) Calibration STARTED. Tcoil start: {start_tcoil}")
                elif status == 'calibrating':
                    send_node_command(mac, "actiondenied:busy")
                    print(f"Node {mac} Action denied. Device is currently busy.")
                else:
                    send_node_command(mac, "calibrationfailack")
                    print(f"Node {mac} Calibration already done or not needed.")

        elif event_type == "calibration_progress":
            if status == 'calibrating':
                try:
                    t3 = float(data.get("t3", 0))
                    base_t3 = float(data.get("base_t3", 0))
                    tracker = CALIBRATION_TRACKER.get(appliance_id, {})
                    if tracker.get('start_tcoil') is None or base_t3 != 0:
                        tracker['start_tcoil'] = base_t3 if base_t3 != 0 else t3
                    tracker['current_tcoil'] = t3
                    CALIBRATION_TRACKER[appliance_id] = tracker
                except (ValueError, TypeError):
                    pass

        elif event_type == "calibration_success_request":
            if status == 'calibrating':
                if appliance_id in CALIBRATION_TRACKER:
                    del CALIBRATION_TRACKER[appliance_id]
                base_data = data.get("base")
                final_data = data.get("final")
                if base_data and final_data:
                    try:
                        t1_delta = abs(float(base_data.get("t1", 0)) - float(final_data.get("t1", 0)))
                        t2_delta = abs(float(base_data.get("t2", 0)) - float(final_data.get("t2", 0)))
                        t3_delta = abs(float(base_data.get("t3", 0)) - float(final_data.get("t3", 0)))
                    except (ValueError, TypeError):
                        t1_delta = t2_delta = t3_delta = 0.0
                    if t3_delta < 7.5 or t2_delta < 2.5 or t1_delta < 2.5:
                        cur.execute("UPDATE appliances SET operational_status='calibration_needed' WHERE id=%s", (appliance_id,))
                        conn.commit()
                        send_node_command(mac, "calibrationfailack")
                        print(f"Node {mac} Calibration REJECTED. Sensors didn't drop enough.")
                    else:
                        def safe_polyfit(x1, x2, y1, y2):
                            try:
                                x1, x2, y1, y2 = float(x1), float(x2), float(y1), float(y2)
                                if abs(x1 - x2) < 0.1: return 1.0, 0.0
                                coeffs = np.polyfit([x1, x2], [y1, y2], 1)
                                return float(coeffs[0]), float(coeffs[1])
                            except Exception:
                                return 1.0, 0.0
                        t1_m, t1_c = safe_polyfit(base_data.get("t1", 0), final_data.get("t1", 0), base_data.get("t3", 0), final_data.get("t3", 0))
                        t2_m, t2_c = safe_polyfit(base_data.get("t2", 0), final_data.get("t2", 0), base_data.get("t3", 0), final_data.get("t3", 0))
                        h2_m, h2_c = safe_polyfit(base_data.get("h2", 0), final_data.get("h2", 0), base_data.get("h1", 0), final_data.get("h1", 0))
                        cur.execute("""
                            UPDATE appliances SET
                                treturn_slope=%s, treturn_intercept=%s,
                                rhreturn_slope=1.0, rhreturn_intercept=0.0,
                                tsupply_slope=%s, tsupply_intercept=%s,
                                rhsupply_slope=%s, rhsupply_intercept=%s,
                                tcoil_slope=1.0, tcoil_offset=0.0,
                                operational_status='normal'
                            WHERE id=%s AND operational_status='calibrating'
                        """, (t1_m, t1_c, t2_m, t2_c, h2_m, h2_c, appliance_id))
                        conn.commit()
                        send_node_command(mac, "calibrationsuccessack")
                        print(f"Node {mac} Multi-Point Linear Calibration SUCCESS.")
                else:
                    send_node_command(mac, "calibrationfailack")

        elif event_type == "calibration_fail_request":
            if status == 'calibrating':
                if appliance_id in CALIBRATION_TRACKER:
                    del CALIBRATION_TRACKER[appliance_id]
                cur.execute("UPDATE appliances SET operational_status = 'calibration_needed' WHERE id = %s", (appliance_id,))
                conn.commit()
                send_node_command(mac, "calibrationfailack")

        elif event_type == "maintenance_request":
            if "Dryer" in app_type or status == 'normal':
                cur.execute("INSERT INTO sensor_events (sensor_node_mac, event_type, timestamp) VALUES (%s,%s,%s)",
                            (mac, 'maintenance', datetime.now()))
                conn.commit()
                send_node_command(mac, "maintenanceack")
                print(f"Node {mac} maintenance logged.")
            else:
                send_node_command(mac, "maintenancedenied")
                print(f"Node {mac} maintenance denied: status={status}")

        cur.close()
        release_conn(conn)
    except Exception as e:
        print(f"Error handling event: {e}")

def on_mqtt_message(client, userdata, msg):
    try:
        payload = msg.payload.decode()
        topic = msg.topic
        parts = topic.split("/")
        if len(parts) < 4: return
        mac = parts[2]
        if "events" in topic:
            handle_node_events(mac, payload)
            return
        elif "telemetry" in topic:
            safe_str = payload.replace('nan', 'null')
            data = json.loads(safe_str)
            now_utc = datetime.now(timezone.utc)
            if "agoms" in data:
                actual_time = now_utc - timedelta(milliseconds=max(0, int(data["agoms"])))
            elif "ago_ms" in data:
                actual_time = now_utc - timedelta(milliseconds=max(0, int(data["ago_ms"])))
            else:
                actual_time = now_utc - timedelta(seconds=max(0, int(data.get("ago", 0))))
            if actual_time > now_utc + timedelta(minutes=1):
                actual_time = now_utc
            if mac in DEDUPE_CACHE:
                diff = abs((actual_time - DEDUPE_CACHE[mac]).total_seconds())
                if diff < 1.0:
                    return
            DEDUPE_CACHE[mac] = actual_time

            conn = get_conn()
            if conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT sn.id, sn.appliance_id, a.type, a.operational_status
                    FROM sensor_nodes sn
                    LEFT JOIN appliances a ON sn.appliance_id = a.id
                    WHERE sn.mac_address = %s
                """, (mac,))
                row = cur.fetchone()
                if not row:
                    cur.execute("INSERT INTO sensor_nodes (mac_address, status, last_seen) VALUES (%s, 'unpaired', NOW()) RETURNING id", (mac,))
                    sensor_node_id = cur.fetchone()[0]
                    appliance_id = None
                    appliance_type = None
                    conn.commit()
                else:
                    sensor_node_id, appliance_id, appliance_type, _ = row
                    cur.execute("UPDATE sensor_nodes SET last_seen = NOW() WHERE id = %s", (sensor_node_id,))
                    conn.commit()

                # Current is computed by the sensor node using backend-provided CF/deductor
                final_amps = max(0.0, float(data.get("CurrentA", 0.0) or 0.0))

                status_field = data.get("status", "running")
                is_running = (status_field == "running")

                reading_values = {}
                if appliance_type:
                    # Always insert readings (both running and idle)
                    if "Dryer" in appliance_type:
                        tex = data.get("BME280Temp")
                        rhex = data.get("BME280Hum")
                        pres = data.get("BME280Pres")
                        cur.execute("""
                            INSERT INTO dryer_readings (sensor_node_id, texhaust, rh_exhaust, pressure, imotor, time)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (sensor_node_id, tex, rhex, pres, final_amps, actual_time))
                        reading_values = {'texhaust': tex, 'rhexhaust': rhex, 'pressure': pres, 'current': final_amps}
                    else:
                        t1 = data.get("DHT1Temp")
                        h1 = data.get("DHT1Hum")
                        t2 = data.get("DHT2Temp")
                        h2 = data.get("DHT2Hum")
                        t3 = data.get("DS18B20Temp")
                        cur.execute("""
                            INSERT INTO hvac_readings (sensor_node_id, treturn, rhreturn, tsupply, rhsupply, tcoil, icompressor, time)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (sensor_node_id, t1, h1, t2, h2, t3, final_amps, actual_time))
                        # Compute deltas for SPC alert checking
                        cal = get_appliance_calibration(appliance_id)
                        t1c = apply_calibration(t1, cal['t1_m'], cal['t1_c'])
                        t2c = apply_calibration(t2, cal['t2_m'], cal['t2_c'])
                        h1c = apply_calibration(h1, cal['h1_m'], cal['h1_c'])
                        h2c = apply_calibration(h2, cal['h2_m'], cal['h2_c'])
                        reading_values = {
                            'deltat': abs(t1c - t2c),
                            'deltarh': abs(h1c - h2c),
                            'tcoil': apply_calibration(t3, cal['tcoil_m'], cal['tcoil_c']),
                            'rhreturn': h1,
                            'rhsupply': h2,
                            'current': final_amps,
                            'treturn': t1c,
                            'tsupply': t2c
                        }
                    cur.execute("UPDATE sensor_nodes SET last_seen = NOW() WHERE id = %s", (sensor_node_id,))
                    conn.commit()

                    # --- Real-time fault alert checking (only on running data) ---
                    if appliance_id and reading_values and final_amps >= 0.25:
                        check_fault_alerts(appliance_id, reading_values, appliance_type, actual_time)
                else:
                    UNPAIRED_CACHE[sensor_node_id] = {
                        "data": data,
                        "amps": final_amps,
                        "time": actual_time,
                        "mac": mac
                    }
                cur.close()
                release_conn(conn)
    except Exception as e:
        print(f"MQTT Message Error: {e}")

mqtt_client.on_connect = on_mqtt_connect
mqtt_client.on_message = on_mqtt_message

# --- AUTH ROUTES ---
@app.route('/')
def landing():
    return redirect(url_for('dashboard')) if current_user.is_authenticated else redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html',
                           user=current_user,
                           appliances=get_appliances_for_user(current_user.id),
                           unpaired_nodes=unpaired_nodes(),
                           all_nodes=get_all_nodes_for_user(current_user.id))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'GET':
        return render_template('signup.html')
    name = request.form.get('name')
    email = request.form.get('email')
    password = request.form.get('password')
    if not name or not email or not password:
        flash('All fields are required', 'error')
        return redirect(url_for('signup'))
    hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    conn = get_conn()
    if conn:
        cur = conn.cursor()
        try:
            cur.execute("INSERT INTO users (email, password_hash, name) VALUES (%s, %s, %s) RETURNING id", (email, hashed, name))
            user_id = cur.fetchone()[0]
            conn.commit()
            user = User(user_id, email, name)
            login_user(user)
            return redirect(url_for('dashboard'))
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
            flash('Email already exists', 'error')
        except Exception as e:
            conn.rollback()
            flash(f'Error: {e}', 'error')
        finally:
            cur.close()
            release_conn(conn)
    return redirect(url_for('signup'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')
    email = request.form.get('email')
    password = request.form.get('password')
    conn = get_conn()
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT id, email, name, password_hash FROM users WHERE email = %s", (email,))
        row = cur.fetchone()
        cur.close()
        release_conn(conn)
        if row and bcrypt.checkpw(password.encode('utf-8'), row[3].encode('utf-8')):
            user = User(row[0], row[1], row[2])
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Invalid credentials', 'error')
    return redirect(url_for('login'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('landing'))

# --- PAIRING ROUTES ---
@app.route('/devices/pair', methods=['POST'])
@login_required
def pair_device():
    name = request.form.get('name')
    dev_type = request.form.get('type')
    node_id = request.form.get('node_id')
    sub_type = request.form.get('sub_type', 'noninverter')
    if not name or not dev_type or not node_id:
        flash('All fields are required', 'error')
        return redirect(url_for('dashboard'))
    conn = get_conn()
    cur = conn.cursor()
    try:
        initial_status = 'normal' if "Dryer" in dev_type else 'calibration_needed'
        cf = 33.0 if "Dryer" in dev_type else 11.0
        deductor = 0.111 if "Dryer" in dev_type else 0.033
        cur.execute("""
            INSERT INTO appliances (user_id, name, type, location, brand, operational_status, sub_type, cf, deductor)
            VALUES (%s, %s, %s, 'Home', 'Generic', %s, %s, %s, %s) RETURNING id
        """, (current_user.id, name, dev_type, initial_status, sub_type, cf, deductor))
        appliance_id = cur.fetchone()[0]
        cur.execute("UPDATE sensor_nodes SET appliance_id = %s, status = 'paired' WHERE id = %s RETURNING mac_address", (appliance_id, node_id))
        mac = cur.fetchone()[0]
        if int(node_id) in UNPAIRED_CACHE:
            del UNPAIRED_CACHE[int(node_id)]
        conn.commit()
        cmd = "settype:dryer" if "Dryer" in dev_type else "settype:hvac"
        send_node_command(mac, cmd)
        if "Dryer" in dev_type:
            send_node_command(mac, "restore:normal")
            flash(f'{name} added! Device is ready. Configure baseline to enable alerts.', 'success')
        else:
            send_node_command(mac, "restore:calibrationneeded")
            flash(f'{name} added! Calibrate sensors, then configure baseline.', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error pairing device: {e}', 'error')
    finally:
        cur.close()
        release_conn(conn)
    return redirect(url_for('dashboard'))

@app.route('/devices/<int:appliance_id>/forget', methods=['POST'])
@login_required
def forget_device(appliance_id):
    conn = get_conn()
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s", (appliance_id,))
        mac_row = cur.fetchone()
        cur.execute("UPDATE sensor_nodes SET appliance_id = NULL, status = 'unpaired' WHERE appliance_id = %s", (appliance_id,))
        cur.execute("DELETE FROM appliances WHERE id = %s", (appliance_id,))
        conn.commit()
        cur.close()
        release_conn(conn)
        if mac_row:
            send_node_command(mac_row[0], "settype:unpaired")
            print(f"Device forgotten: sent settype:unpaired to {mac_row[0]}")
        flash('Device forgotten.', 'success')
    # Clean up in-memory trackers to prevent leaks
    for tracker in (DRYER_CYCLE_STATS, HVAC_CYCLE_TRACKER, FAULT_ALERT_TRACKER,
                    FAULT_ALERT_COOLDOWN,
                    CALIBRATION_TRACKER):
        tracker.pop(appliance_id, None)
    return redirect(url_for('dashboard'))

# --- API: UNPAIRED NODES ---
@app.route('/api/unpaired_nodes')
@login_required
def api_unpaired():
    return jsonify(unpaired_nodes())

@app.route('/api/node/<int:node_id>/latest')
@login_required
def api_node_latest(node_id):
    if node_id in UNPAIRED_CACHE:
        cache = UNPAIRED_CACHE[node_id]
        d = cache["data"]
        if "BME280Temp" in d:
            return jsonify({
                "time": cache["time"].isoformat(),
                "Texhaust": d.get("BME280Temp", 0),
                "RHexhaust": d.get("BME280Hum", 0),
                "Pressure": round(d.get("BME280Pres", 0), 2),
                "Imotor": cache.get("amps", 0),
                "cal_state": d.get("cal_state", "idle")
            })
        else:
            return jsonify({
                "time": cache["time"].isoformat(),
                "Treturn": d.get("DHT1Temp", 0),
                "RHreturn": d.get("DHT1Hum", 0),
                "Tsupply": d.get("DHT2Temp", 0),
                "RHsupply": d.get("DHT2Hum", 0),
                "Tcoil": d.get("DS18B20Temp", 0),
                "Icompressor": cache.get("amps", 0),
                "cal_state": d.get("cal_state", "idle")
            })
    return jsonify({'error': 'Node not in cache'}), 404

# --- API: DEVICE DATA ---
@app.route('/api/device/<int:appliance_id>/latest')
@login_required
def api_device_latest(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("""
        SELECT a.operational_status, a.type, sn.last_seen, a.baseline_configured
        FROM appliances a
        LEFT JOIN sensor_nodes sn ON sn.appliance_id = a.id
        WHERE a.id = %s
    """, (appliance_id,))
    status_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not status_row: return jsonify({'error': 'not found'}), 404

    operational_status = status_row[0]
    dev_type = status_row[1]
    last_seen = status_row[2]
    baseline_configured = status_row[3]
    is_calibrated = operational_status not in ['calibration_needed', 'calibrating']
    alert_status = get_appliance_alert_status(appliance_id)

    offline_threshold_seconds = 660
    is_offline = False
    ever_connected = last_seen is not None
    now = datetime.now(timezone.utc)
    if last_seen:
        if last_seen.tzinfo is None:
            last_seen = last_seen.replace(tzinfo=timezone.utc)
        is_offline = (now - last_seen).total_seconds() > offline_threshold_seconds
    else:
        is_offline = True

    if "Dryer" not in dev_type and operational_status in ['calibration_needed', 'calibrating']:
        return jsonify({
            'error': 'Sensor not calibrated',
            'calibrated': False,
            'status': operational_status,
            'type': dev_type,
            'running_status': 'idle',
            'is_offline': is_offline,
            'ever_connected': ever_connected,
            'has_data': False,
            'baseline_configured': baseline_configured,
            'alert_status': alert_status
        }), 200

    row_data, dev_type = latest_row_for_appliance(appliance_id)
    cal = get_appliance_calibration(appliance_id)

    if not row_data or 't1_m' not in cal:
        return jsonify({
            'type': dev_type,
            'calibrated': is_calibrated and ('t1_m' in cal),
            'status': operational_status,
            'running_status': 'idle',
            'is_offline': is_offline,
            'ever_connected': ever_connected,
            'has_data': False,
            'baseline_configured': baseline_configured,
            'alert_status': alert_status
        }), 200

    time_val = row_data[0]
    stale_threshold_seconds = 60
    if time_val.tzinfo is None:
        time_val = time_val.replace(tzinfo=timezone.utc)
    is_stale = time_val > now or (now - time_val).total_seconds() > stale_threshold_seconds

    if "Dryer" in dev_type:
        imotor = max(0.0, row_data[4] or 0)
        running_status = 'idle' if is_stale else ('running' if imotor >= 0.25 else 'idle')
        return jsonify({
            'time': time_val.isoformat(),
            'Texhaust': apply_calibration(row_data[1], cal['t1_m'], cal['t1_c']),
            'RHexhaust': apply_calibration(row_data[2], cal['h1_m'], cal['h1_c']),
            'Pressure': round(row_data[3] or 0.0, 2),
            'Imotor': imotor,
            'type': dev_type,
            'calibrated': True,
            'status': operational_status,
            'running_status': running_status,
            'data_stale': is_stale,
            'is_offline': is_offline,
            'ever_connected': ever_connected,
            'has_data': True,
            'baseline_configured': baseline_configured,
            'alert_status': alert_status
        })
    else:
        icomp = row_data[6] or 0
        running_status = 'idle' if is_stale else ('running' if icomp >= 0.25 else 'idle')
        t1c = apply_calibration(row_data[1], cal['t1_m'], cal['t1_c'])
        t2c = apply_calibration(row_data[3], cal['t2_m'], cal['t2_c'])
        h1c = apply_calibration(row_data[2], cal['h1_m'], cal['h1_c'])
        h2c = apply_calibration(row_data[4], cal['h2_m'], cal['h2_c'])
        return jsonify({
            'time': time_val.isoformat(),
            'Treturn': t1c,
            'RHreturn': row_data[2],
            'Tsupply': t2c,
            'RHsupply': row_data[4],
            'Tcoil': apply_calibration(row_data[5], cal['tcoil_m'], cal['tcoil_c']),
            'Icompressor': icomp,
            'DeltaT': round(abs(t1c - t2c), 2),
            'DeltaRH': round(abs(h1c - h2c), 2),
            'type': dev_type,
            'calibrated': is_calibrated,
            'status': operational_status,
            'running_status': running_status,
            'data_stale': is_stale,
            'is_offline': is_offline,
            'ever_connected': ever_connected,
            'has_data': True,
            'baseline_configured': baseline_configured,
            'alert_status': alert_status
        })

@app.route('/api/device/<int:appliance_id>/latest_n')
@login_required
def api_device_latest_n(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("SELECT operational_status, type FROM appliances WHERE id = %s", (appliance_id,))
    status_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not status_row or ("Dryer" not in status_row[1] and status_row[0] in ['calibration_needed', 'calibrating']):
        return jsonify([]), 200
    limit = int(request.args.get('limit', 1080))
    start = request.args.get('start')
    end = request.args.get('end')
    filtered = request.args.get('filtered', 'true').lower() != 'false'
    rows, dev_type = latest_n_rows_for_appliance(appliance_id, limit, start, end, filtered)
    cal = get_appliance_calibration(appliance_id)
    if 't1_m' not in cal: return jsonify([])
    result = []
    for r in rows:
        time_val = r[0]
        if "Dryer" in dev_type:
            result.append({
                'time': time_val.isoformat(),
                'Texhaust': apply_calibration(r[1], cal['t1_m'], cal['t1_c']),
                'RHexhaust': apply_calibration(r[2], cal['h1_m'], cal['h1_c']),
                'Pressure': round(r[3] or 0.0, 2),
                'Imotor': max(0.0, r[4] or 0)
            })
        else:
            t1c = apply_calibration(r[1], cal['t1_m'], cal['t1_c'])
            t2c = apply_calibration(r[3], cal['t2_m'], cal['t2_c'])
            h1c = apply_calibration(r[2], cal['h1_m'], cal['h1_c'])
            h2c = apply_calibration(r[4], cal['h2_m'], cal['h2_c'])
            result.append({
                'time': time_val.isoformat(),
                'Treturn': t1c,
                'RHreturn': r[2],
                'Tsupply': t2c,
                'RHsupply': r[4],
                'Tcoil': apply_calibration(r[5], cal['tcoil_m'], cal['tcoil_c']),
                'Icompressor': r[6] or 0,
                'DeltaT': round(abs(t1c - t2c), 2),
                'DeltaRH': round(abs(h1c - h2c), 2)
            })
    return jsonify(result)

@app.route('/api/device/<int:appliance_id>/table_data')
@login_required
def get_table_data(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("SELECT operational_status, type FROM appliances WHERE id = %s", (appliance_id,))
    status_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not status_row or ("Dryer" not in status_row[1] and status_row[0] in ['calibration_needed', 'calibrating']):
        return jsonify([]), 200
    filtered = request.args.get('filtered', 'true').lower() != 'false'
    rows, dev_type = latest_n_rows_for_appliance(appliance_id, 120, filtered=filtered)
    cal = get_appliance_calibration(appliance_id)
    if 't1_m' not in cal: return jsonify([])
    result = []
    for r in reversed(rows):
        time_val = r[0]
        t1 = apply_calibration(r[1], cal['t1_m'], cal['t1_c'])
        h1 = apply_calibration(r[2], cal['h1_m'], cal['h1_c'])
        if "Dryer" in dev_type:
            result.append({
                'time': time_val.isoformat(),
                'Texhaust': round(t1, 2),
                'RHexhaust': round(h1, 2),
                'Pressure': round(r[3] or 0.0, 2),
                'Imotor': round(max(0.0, r[4] or 0), 2)
            })
        else:
            t2 = apply_calibration(r[3], cal['t2_m'], cal['t2_c'])
            h2 = apply_calibration(r[4], cal['h2_m'], cal['h2_c'])
            t3 = apply_calibration(r[5], cal['tcoil_m'], cal['tcoil_c'])
            amp = r[6] or 0
            result.append({
                'time': time_val.isoformat(),
                'Treturn': round(t1, 2),
                'Tsupply': round(t2, 2),
                'Tcoil': round(t3, 2),
                'RHreturn': round(r[2], 2),
                'RHsupply': round(r[4], 2),
                'Icompressor': round(amp, 2),
                'DeltaT': round(abs(t1 - t2), 2),
                'DeltaRH': round(abs(h1 - h2), 2)
            })
    return jsonify(result)

@app.route('/api/device/<int:appliance_id>/maintenance_logs')
@login_required
def get_maintenance_logs(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s LIMIT 1", (appliance_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); release_conn(conn)
        return jsonify([]), 200
    cur.execute("""
        SELECT DISTINCT se.timestamp
        FROM sensor_events se
        JOIN sensor_nodes sn ON se.sensor_node_mac = sn.mac_address
        JOIN appliances a ON sn.appliance_id = a.id
        WHERE se.sensor_node_mac = %s
          AND se.event_type = 'maintenance'
          AND se.timestamp >= a.created_at
        ORDER BY se.timestamp DESC
    """, (row[0],))
    rows = cur.fetchall()
    cur.close(); release_conn(conn)
    return jsonify([r[0].isoformat() for r in rows])

# --- API: EXPORT EXCEL ---
@app.route('/api/device/<int:appliance_id>/export_excel')
@login_required
def export_excel(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db error'}), 500
    cur = conn.cursor()
    cur.execute("SELECT name, type, sub_type, operational_status, created_at FROM appliances WHERE id = %s", (appliance_id,))
    app_info = cur.fetchone()
    if not app_info:
        cur.close(); release_conn(conn)
        return jsonify({'error': 'Not found'}), 404
    app_name, dev_type, sub_type, status, created_at = app_info
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    filtered = request.args.get('filtered', 'true').lower() != 'false'
    query_params = [appliance_id]
    date_filter = ""
    if start_date:
        date_filter += " AND r.time >= %s"
        query_params.append(start_date)
    if end_date:
        # Pad end by 1s to match dryer_analytics / hvac_analytics inclusive behavior
        try:
            end_dt = datetime.fromisoformat(end_date)
            end_date = (end_dt + timedelta(seconds=1)).isoformat()
        except ValueError:
            pass
        date_filter += " AND r.time <= %s"
        query_params.append(end_date)
    current_filter = " AND r.imotor >= 0.25" if (filtered and "Dryer" in dev_type) else ""
    current_filter_hvac = " AND r.icompressor >= 0.25" if (filtered and "Dryer" not in dev_type) else ""
    if "Dryer" in dev_type:
        query = f"""
            SELECT r.time, r.texhaust, r.rh_exhaust, r.pressure, r.imotor
            FROM dryer_readings r
            JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND r.time >= a.created_at {date_filter}{current_filter}
            ORDER BY r.time ASC
        """
    else:
        query = f"""
            SELECT r.time, r.treturn, r.rhreturn, r.tsupply, r.rhsupply, r.tcoil, r.icompressor
            FROM hvac_readings r
            JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND r.time >= a.created_at {date_filter}{current_filter_hvac}
            ORDER BY r.time ASC
        """
    cur.execute(query, tuple(query_params))
    readings = cur.fetchall()
    cal = get_appliance_calibration(appliance_id)
    cur.close(); release_conn(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sensor Data"
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Device: {app_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A2:D2')
    if "Dryer" in dev_type:
        ws['A2'] = f"Type: {dev_type}"
    else:
        ws['A2'] = f"Type: {dev_type} ({sub_type if sub_type else ''})"
    ws.merge_cells('A3:D3')
    ws['A3'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A4:D4')
    ws['A4'] = f"Data Points: {len(readings)}"
    row_idx = 6
    if "Dryer" in dev_type:
        headers = ["Timestamp", "Exhaust Temp (°C)", "Exhaust RH (%)", "Pressure (hPa)", "Current (A)"]
    else:
        headers = ["Timestamp", "Return Temp (°C)", "Supply Temp (°C)", "Coil Temp (°C)", "Return RH (%)", "Supply RH (%)", "Current (A)", "Delta-T (°C)", "Delta-RH (%)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    if not readings:
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(headers))
        no_data_cell = ws.cell(row=7, column=1, value="No sensor data available for this device in the selected range.")
        no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_data_cell.font = Font(italic=True, color="999999")
    else:
        for r in readings:
            row_idx += 1
            t1 = apply_calibration(r[1], cal['t1_m'], cal['t1_c'])
            h1 = apply_calibration(r[2], cal['h1_m'], cal['h1_c'])
            ws.cell(row=row_idx, column=1, value=r[0].strftime('%Y-%m-%d %H:%M:%S'))
            if "Dryer" in dev_type:
                ws.cell(row=row_idx, column=2, value=round(t1, 2))
                ws.cell(row=row_idx, column=3, value=round(h1, 2))
                ws.cell(row=row_idx, column=4, value=round(r[3] or 0.0, 2))
                ws.cell(row=row_idx, column=5, value=round(r[4] or 0, 2))
            else:
                t2 = apply_calibration(r[3], cal['t2_m'], cal['t2_c'])
                h2 = apply_calibration(r[4], cal['h2_m'], cal['h2_c'])
                t3 = apply_calibration(r[5], cal['tcoil_m'], cal['tcoil_c'])
                amp = r[6] or 0
                ws.cell(row=row_idx, column=2, value=round(t1, 2))
                ws.cell(row=row_idx, column=3, value=round(t2, 2))
                ws.cell(row=row_idx, column=4, value=round(t3, 2))
                ws.cell(row=row_idx, column=5, value=round(r[2], 2))
                ws.cell(row=row_idx, column=6, value=round(r[4], 2))
                ws.cell(row=row_idx, column=7, value=round(amp, 2))
                ws.cell(row=row_idx, column=8, value=round(abs(t1 - t2), 2))
                ws.cell(row=row_idx, column=9, value=round(abs(h1 - h2), 2))
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx2 in range(6, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx2, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    conn = get_conn()
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s LIMIT 1", (appliance_id,))
        mac_row = cur.fetchone()
        maintenance_rows = []
        if mac_row:
            cur.execute("""
                SELECT timestamp FROM sensor_events
                WHERE sensor_node_mac = %s AND event_type = 'maintenance'
                  AND timestamp >= (SELECT created_at FROM appliances WHERE id = %s)
                ORDER BY timestamp DESC
            """, (mac_row[0], appliance_id))
            maintenance_rows = cur.fetchall()
        cur.close()
        release_conn(conn)
        ws2 = wb.create_sheet(title="Maintenance Log")
        ws2['A1'] = f"Device: {app_name}"
        ws2['A1'].font = Font(size=14, bold=True)
        ws2['A3'] = "Maintenance Date"
        ws2['A3'].fill = header_fill
        ws2['A3'].font = header_font
        ws2['A3'].alignment = Alignment(horizontal="center")
        row_idx2 = 3
        for r_maint in maintenance_rows:
            row_idx2 += 1
            ws2.cell(row=row_idx2, column=1, value=r_maint[0].strftime('%Y-%m-%d %H:%M:%S'))
        ws2.column_dimensions['A'].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{app_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# --- API: CALIBRATION PROGRESS ---
@app.route('/api/device/<int:appliance_id>/calibration_progress')
@login_required
def api_calibration_progress(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT operational_status, type, created_at FROM appliances WHERE id = %s", (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not row: return jsonify({'error': 'not found'}), 404
    status, dev_type, created_at = row[0], row[1], row[2]
    if status != 'calibrating':
        return jsonify({'error': 'Not calibrating'}), 400
    tracker = CALIBRATION_TRACKER.get(appliance_id, {})
    start_tcoil = tracker.get('start_tcoil')
    current_tcoil = tracker.get('current_tcoil')

    # Fallback to DB if tracker hasn't been populated yet (old firmware or edge case)
    if start_tcoil is None or current_tcoil is None:
        conn = get_conn()
        if conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT sr.tcoil FROM hvac_readings sr
                JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND sr.time >= a.created_at
                ORDER BY sr.time DESC LIMIT 1
            """, (appliance_id,))
            tcoil_row = cur.fetchone()
            cur.close()
            release_conn(conn)
            db_tcoil = float(tcoil_row[0]) if tcoil_row and tcoil_row[0] is not None else None
            if start_tcoil is None:
                start_tcoil = db_tcoil if db_tcoil is not None else 25.0
            if current_tcoil is None:
                current_tcoil = db_tcoil if db_tcoil is not None else start_tcoil
        else:
            if start_tcoil is None:
                start_tcoil = 25.0
            if current_tcoil is None:
                current_tcoil = start_tcoil

    drop = start_tcoil - current_tcoil
    return jsonify({
        'start_tcoil': round(start_tcoil, 2),
        'current_tcoil': round(current_tcoil, 2),
        'drop': round(drop, 2),
        'target': 8.0,
        'progress_pct': min(100, max(0, round((drop / 8.0) * 100, 1)))
    })

# --- API: SPC LIMITS (reads from spc_manual_baselines) ---
@app.route('/api/device/<int:appliance_id>/spc_limits')
@login_required
def api_spc_limits(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT type, sub_type, operational_status, baseline_configured FROM appliances WHERE id = %s", (appliance_id,))
    db_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not db_row: return jsonify({'error': 'not found'}), 404
    dev_type, subtype, status, baseline_configured = db_row[0], db_row[1], db_row[2], db_row[3]
    baselines = get_spc_baselines(appliance_id)
    is_inverter = (subtype == 'inverter')

    def make_limit(metric_name):
        if metric_name in baselines:
            b = baselines[metric_name]
            return {"mean": b['mean'], "ucl": b['ucl'], "lcl": b['lcl']}
        return {"mean": 0, "ucl": 0, "lcl": 0}

    expected = DRYER_METRICS if "Dryer" in dev_type else HVAC_METRICS
    result = {
        "type": dev_type,
        "subtype": subtype,
        "status": status,
        "alert_status": get_appliance_alert_status(appliance_id),
        "baseline_configured": baseline_configured,
    }
    for m in expected:
        result[m] = make_limit(m)
    return jsonify(result)

# --- API: BASELINE ANALYSIS (reads from spc_manual_baselines) ---
@app.route('/api/device/<int:appliance_id>/baseline_analysis')
@login_required
def api_baseline_analysis(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT type, sub_type, operational_status, baseline_configured FROM appliances WHERE id = %s", (appliance_id,))
    db_row = cur.fetchone()
    # Get most recent baseline update timestamp
    cur.execute("SELECT MAX(updated_at) FROM spc_manual_baselines WHERE appliance_id = %s", (appliance_id,))
    ts_row = cur.fetchone()
    baseline_set_at = ts_row[0].isoformat() if ts_row and ts_row[0] else None
    cur.close()
    release_conn(conn)
    if not db_row: return jsonify({'error': 'not found'}), 404
    dev_type, subtype, status, baseline_configured = db_row[0], db_row[1], db_row[2], db_row[3]
    baselines = get_spc_baselines(appliance_id)

    def metric_info(name):
        if name in baselines:
            b = baselines[name]
            return {"mean": b['mean'], "ucl": b['ucl'], "lcl": b['lcl']}
        return None

    expected = DRYER_METRICS if "Dryer" in dev_type else HVAC_METRICS
    result = {
        "type": dev_type,
        "status": status,
        "baseline_configured": baseline_configured,
        "baseline_set_at": baseline_set_at,
    }
    for m in expected:
        result[m] = metric_info(m)
    return jsonify(result)

# --- API: BASELINE CONFIG (NEW) ---
@app.route('/api/device/<int:appliance_id>/baseline_config', methods=['GET', 'POST', 'DELETE'])
@login_required
def api_baseline_config(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT type, operational_status FROM appliances WHERE id = %s", (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not row: return jsonify({'error': 'not found'}), 404
    dev_type, status = row[0], row[1]

    if request.method == 'GET':
        baselines = get_spc_baselines(appliance_id)
        expected = DRYER_METRICS if "Dryer" in dev_type else HVAC_METRICS
        result = {}
        for m in expected:
            if m in baselines:
                result[m] = baselines[m]
            else:
                result[m] = {'ucl': '', 'lcl': '', 'mean': ''}
        return jsonify({'type': dev_type, 'metrics': result})

    if request.method == 'DELETE':
        conn = get_conn()
        if not conn: return jsonify({'error': 'db'}), 500
        cur = conn.cursor()
        try:
            cur.execute("DELETE FROM spc_manual_baselines WHERE appliance_id = %s", (appliance_id,))
            cur.execute("UPDATE appliances SET baseline_configured = FALSE WHERE id = %s", (appliance_id,))
            conn.commit()
            return jsonify({'success': True, 'message': 'Baseline removed'})
        except Exception as e:
            conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            cur.close()
            release_conn(conn)

    # POST
    data = request.get_json() or {}
    metrics = data.get('metrics', {})
    expected = DRYER_METRICS if "Dryer" in dev_type else HVAC_METRICS
    baselines_to_save = {}
    for m in expected:
        if m in metrics:
            try:
                ucl_raw = metrics[m].get('ucl')
                lcl_raw = metrics[m].get('lcl')
                mean_raw = metrics[m].get('mean')
                # Auto-derive UCL/LCL for dryer current if mean is provided but UCL/LCL are empty
                if "Dryer" in dev_type and m == 'current' and mean_raw:
                    mean_val = float(mean_raw)
                    if (not ucl_raw or str(ucl_raw).strip() == '') and (not lcl_raw or str(lcl_raw).strip() == ''):
                        ucl = round(mean_val * 1.20, 3)
                        lcl = round(mean_val * 0.80, 3)
                    else:
                        ucl = float(ucl_raw)
                        lcl = float(lcl_raw)
                else:
                    ucl = float(ucl_raw)
                    lcl = float(lcl_raw)
                if ucl <= lcl:
                    return jsonify({'error': f'UCL must be greater than LCL for {m}'}), 400
                baselines_to_save[m] = {'ucl': ucl, 'lcl': lcl}
            except (ValueError, TypeError):
                return jsonify({'error': f'Invalid values for {m}'}), 400
    if not baselines_to_save:
        return jsonify({'error': 'No valid baseline data provided'}), 400
    success, msg = save_spc_baselines(appliance_id, baselines_to_save)
    if success:
        notify_node_baseline_set(appliance_id)
        return jsonify({'success': True, 'message': msg})
    return jsonify({'error': msg}), 500

# --- API: SENSOR CONFIG (CF / Deductor) ---
@app.route('/api/device/<int:appliance_id>/sensor_config', methods=['GET', 'POST'])
@login_required
def api_sensor_config(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db error'}), 500
    cur = conn.cursor()
    cur.execute("SELECT user_id, type, cf, deductor FROM appliances WHERE id = %s", (appliance_id,))
    row = cur.fetchone()
    if not row or row[0] != current_user.id:
        cur.close(); release_conn(conn)
        return jsonify({'error': 'unauthorized'}), 403
    app_type = row[1]
    if request.method == 'GET':
        cf = row[2] if row[2] is not None else (33.0 if "Dryer" in app_type else 11.0)
        deductor = row[3] if row[3] is not None else (0.111 if "Dryer" in app_type else 0.033)
        cur.close(); release_conn(conn)
        return jsonify({'cf': cf, 'deductor': deductor})
    data = request.get_json() or {}
    cf = data.get('cf')
    deductor = data.get('deductor')
    if cf is None or deductor is None:
        cur.close(); release_conn(conn)
        return jsonify({'error': 'cf and deductor required'}), 400
    cur.execute("UPDATE appliances SET cf = %s, deductor = %s WHERE id = %s", (cf, deductor, appliance_id))
    conn.commit()
    cur.close(); release_conn(conn)
    return jsonify({'success': True})

# --- API: THRESHOLDS ---
@app.route('/api/device/<int:appliance_id>/thresholds', methods=['GET', 'POST'])
@login_required
def api_thresholds(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    if request.method == 'GET':
        cur = conn.cursor()
        cur.execute("SELECT alert_rhexhaust_threshold, alert_enabled FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        cur.close()
        release_conn(conn)
        if not row: return jsonify({'error': 'not found'}), 404
        return jsonify({
            "alert_rhexhaust_threshold": float(row[0]) if row[0] else 40.0,
            "alert_enabled": row[1] if row[1] is not None else True
        })
    try:
        cur = conn.cursor()
        data = request.get_json() or {}
        threshold = data.get('alert_rhexhaust_threshold')
        enabled = data.get('alert_enabled')
        updates = []
        params = []
        if threshold is not None:
            updates.append("alert_rhexhaust_threshold = %s")
            params.append(float(threshold))
        if enabled is not None:
            updates.append("alert_enabled = %s")
            params.append(bool(enabled))
        if updates:
            params.append(appliance_id)
            cur.execute(f"UPDATE appliances SET {', '.join(updates)} WHERE id = %s", params)
            conn.commit()
        cur.close()
        release_conn(conn)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- API: ALERTS ---
@app.route('/api/device/<int:appliance_id>/alerts')
@login_required
def api_alerts(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("""
        SELECT a.id, a.alert_type, a.message, a.value, a.threshold, a.severity, a.created_at, a.resolved_at, a.acknowledged
        FROM alerts a
        WHERE a.appliance_id = %s AND a.alert_type LIKE 'fault_%%'
        ORDER BY a.created_at DESC LIMIT 50
    """, (appliance_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    result = []
    for r in rows:
        result.append({
            "id": r[0], "alert_type": r[1], "message": r[2],
            "value": float(r[3]) if r[3] is not None else None,
            "threshold": float(r[4]) if r[4] is not None else None,
            "severity": r[5],
            "created_at": r[6].isoformat() if r[6] else None,
            "resolved_at": r[7].isoformat() if r[7] else None,
            "acknowledged": r[8]
        })
    return jsonify(result)

@app.route('/api/alert/<int:alert_id>/resolve', methods=['POST'])
@login_required
def resolve_alert(alert_id):
    """Resolve a single alert by setting resolved_at = NOW()."""
    conn = get_conn()
    if not conn:
        return jsonify({'error': 'db error'}), 500
    cur = conn.cursor()
    try:
        # Verify the alert belongs to an appliance owned by the current user
        cur.execute("""
            SELECT a.appliance_id, ap.user_id
            FROM alerts a
            JOIN appliances ap ON a.appliance_id = ap.id
            WHERE a.id = %s
        """, (alert_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'alert not found'}), 404
        appliance_id, owner_id = row
        if owner_id != current_user.id:
            return jsonify({'error': 'unauthorized'}), 403
        # Set resolved_at
        cur.execute("UPDATE alerts SET resolved_at = NOW() WHERE id = %s", (alert_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Alert resolved'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        release_conn(conn)

# --- API: HVAC ANALYTICS ---
@app.route('/api/device/<int:appliance_id>/hvac_analytics')
@login_required
def hvac_analytics(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    try:
        cur = conn.cursor()
        cur.execute("SELECT type FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        if not row or "HVAC" not in row[0]:
            return jsonify({'error': 'Not HVAC'}), 400
        start = request.args.get('start')
        end = request.args.get('end')
        # Pad end by 1s to include any milliseconds for inclusive time-range queries
        if end:
            try:
                end_dt = datetime.fromisoformat(end)
                end = (end_dt + timedelta(seconds=1)).isoformat()
            except ValueError:
                pass

        # --- Daily Averages ---
        if start and end:
            cur.execute("""
                SELECT DATE(r.time) as date,
                       AVG(r.treturn), AVG(r.tsupply), AVG(r.tcoil)
                FROM hvac_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.icompressor >= 0.25 AND r.time >= a.created_at AND r.time >= %s AND r.time <= %s
                GROUP BY DATE(r.time)
                ORDER BY DATE(r.time) DESC
                LIMIT 30
            """, (appliance_id, start, end))
        else:
            cur.execute("""
                SELECT DATE(r.time) as date,
                       AVG(r.treturn), AVG(r.tsupply), AVG(r.tcoil)
                FROM hvac_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.icompressor >= 0.25 AND r.time >= a.created_at
                GROUP BY DATE(r.time)
                ORDER BY DATE(r.time) DESC
                LIMIT 30
            """, (appliance_id,))
        daily_rows = cur.fetchall()
        daily_averages = []
        for r in daily_rows:
            daily_averages.append({
                "date": r[0].isoformat(),
                "avg_intake": round(r[1], 2),
                "avg_exit": round(r[2], 2),
                "avg_coil": round(r[3], 2)
            })

        # --- Daily Energy (kWh) ---
        voltage = get_appliance_voltage(appliance_id)
        if start and end:
            cur.execute("""
                SELECT r.time, r.icompressor
                FROM hvac_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.time >= a.created_at AND r.time >= %s AND r.time <= %s
                ORDER BY r.time ASC
            """, (appliance_id, start, end))
        else:
            cur.execute("""
                SELECT r.time, r.icompressor
                FROM hvac_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.time >= a.created_at
                ORDER BY r.time ASC
            """, (appliance_id,))
        readings = cur.fetchall()

        # Group readings by date and compute energy per day
        readings_by_date = defaultdict(list)
        for r in readings:
            date_key = r[0].date()
            readings_by_date[date_key].append(r)

        for day in daily_averages:
            date_key = datetime.fromisoformat(day["date"]).date()
            day_readings = readings_by_date.get(date_key, [])
            day["daily_energy_kwh"] = _compute_daily_energy(day_readings, voltage)

        return jsonify({"daily_averages": daily_averages})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cur.close()
            release_conn(conn)

# --- API: DRYER ANALYTICS ---
@app.route('/api/device/<int:appliance_id>/dryer_analytics')
@login_required
def dryer_analytics(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    try:
        cur = conn.cursor()
        cur.execute("SELECT type FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        if not row or "Dryer" not in row[0]:
            return jsonify({'error': 'Not a dryer'}), 400
        cycle_start = 0.25
        cycle_end = 0.15
        baselines = get_spc_baselines(appliance_id)
        mean_current = baselines.get('current', {}).get('mean', 2.0)
        prominence_threshold = 0.40
        start = request.args.get('start')
        end = request.args.get('end')
        # Pad end by 1s to include any milliseconds for inclusive time-range queries
        if end:
            try:
                end_dt = datetime.fromisoformat(end)
                end = (end_dt + timedelta(seconds=1)).isoformat()
            except ValueError:
                pass
        if start and end:
            cur.execute("""
                SELECT r.time, r.texhaust, r.rh_exhaust, r.pressure, r.imotor
                FROM dryer_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.time >= a.created_at AND r.time >= %s AND r.time <= %s AND r.imotor >= 0.25
                ORDER BY r.time ASC
            """, (appliance_id, start, end))
        else:
            cur.execute("""
                SELECT r.time, r.texhaust, r.rh_exhaust, r.pressure, r.imotor
                FROM dryer_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.time >= a.created_at AND r.imotor >= 0.25
                ORDER BY r.time ASC
            """, (appliance_id,))
        readings = cur.fetchall()
        if not readings:
            return jsonify([])

        # --- DEBUG LOGGING ---

        # ---------------------

        voltage = get_appliance_voltage(appliance_id)
        cycles = []
        in_cycle = False
        current_cycle = {}
        _prev_current = 0.0
        _peak_state = "IDLE"
        _peak_max = 0.0
        _peak_valley = 0.0
        _peak_values = []

        def _confirm_peak():
            nonlocal _peak_state, _peak_max, _peak_valley
            if _peak_state in ("RISING", "FALLING") and _peak_max > 0:
                prominence = _peak_max - _peak_valley
                if prominence >= prominence_threshold and _peak_max > mean_current + 0.15:
                    _peak_values.append(_peak_max)
            _peak_state = "IDLE"
            _peak_max = 0.0
            _peak_valley = 0.0

        for i, r in enumerate(readings):
            time_val, tex, rhex, press, imotor = r
            imotor = float(imotor) if imotor is not None else 0.0
            tex = float(tex) if tex is not None else 0.0
            rhex = float(rhex) if rhex is not None else None
            if in_cycle and i > 0:
                gap = (time_val - readings[i-1][0]).total_seconds()
                if gap > 120:
                    in_cycle = False
                    prev_time = readings[i-1][0]
                    current_cycle["end_time"] = prev_time
                    duration = current_cycle["end_time"] - current_cycle["start_time"]
                    current_cycle["duration_minutes"] = round(duration.total_seconds() / 60, 1)
                    valid_rh = [v for v in current_cycle["_rh_history"] if v is not None]
                    first_6 = valid_rh[:6] if len(valid_rh) > 6 else valid_rh
                    current_cycle["start_rh"] = round(sum(first_6) / len(first_6), 2) if first_6 else None
                    last_6 = valid_rh[-6:] if len(valid_rh) > 6 else valid_rh
                    current_cycle["end_rh_avg"] = round(sum(last_6) / len(last_6), 2) if last_6 else current_cycle.get("start_rh", 0)
                    currents = current_cycle["_currents"]
                    times = current_cycle["_times"]
                    energy_ws = 0.0
                    for j in range(1, len(currents)):
                        dt = (times[j] - times[j-1]).total_seconds()
                        energy_ws += currents[j-1] * voltage * dt
                    current_cycle["energy_kwh"] = round(energy_ws / 3_600_000, 4)
                    _confirm_peak()
                    current_cycle["current_spike_avg"] = round(sum(_peak_values) / len(_peak_values), 2) if _peak_values else 0.0
                    current_cycle["ignition_count"] = len(_peak_values)
                    motor_readings = current_cycle.get("_motor_readings", [])
                    filter_threshold = None
                    if motor_readings:
                        filter_threshold = (sum(motor_readings) / len(motor_readings)) * 1.15
                    median_val = _compute_motor_baseline_median(motor_readings, filter_threshold=filter_threshold)
                    current_cycle["motor_baseline_median"] = round(median_val, 3)
                    del current_cycle["_rh_history"]
                    del current_cycle["_currents"]
                    del current_cycle["_times"]
                    del current_cycle["_motor_readings"]
                    current_cycle["start_time"] = current_cycle["start_time"].isoformat()
                    current_cycle["end_time"] = current_cycle["end_time"].isoformat()
                    cycles.append(current_cycle)

                    current_cycle = {}
                    _prev_current = 0.0
                    _peak_values = []
            if imotor > cycle_start and not in_cycle:
                in_cycle = True
                current_cycle = {
                    "start_time": time_val,
                    "min_temp": tex,
                    "max_temp": tex,
                    "start_rh": None,  # computed at finalization from first 6 readings
                    "_rh_history": [],
                    "_currents": [],
                    "_times": [],
                    "_motor_readings": []
                }
                _prev_current = imotor
                _confirm_peak()
                _peak_values = []
            if in_cycle:
                current_cycle["max_temp"] = max(current_cycle["max_temp"], tex)
                current_cycle["min_temp"] = min(current_cycle["min_temp"], tex)
                if rhex is not None:
                    current_cycle["_rh_history"].append(rhex)
                current_cycle["_currents"].append(imotor)
                current_cycle["_times"].append(time_val)
                current_cycle["_motor_readings"].append(imotor)
                if imotor > _prev_current:
                    if _peak_state in ("IDLE", "FALLING"):
                        if _peak_state == "FALLING":
                            _confirm_peak()
                        _peak_valley = _prev_current
                    _peak_state = "RISING"
                    if imotor > _peak_max:
                        _peak_max = imotor
                elif imotor < _prev_current:
                    if _peak_state == "RISING" and (_peak_max <= 0.1 or imotor < _peak_max - 0.1):
                        _peak_state = "FALLING"
                _prev_current = imotor
            if imotor < cycle_end and in_cycle:
                in_cycle = False
                current_cycle["end_time"] = time_val
                duration = current_cycle["end_time"] - current_cycle["start_time"]
                current_cycle["duration_minutes"] = round(duration.total_seconds() / 60, 1)
                valid_rh = [v for v in current_cycle["_rh_history"] if v is not None]
                first_6 = valid_rh[:6] if len(valid_rh) > 6 else valid_rh
                current_cycle["start_rh"] = round(sum(first_6) / len(first_6), 2) if first_6 else None
                last_6 = valid_rh[-6:] if len(valid_rh) > 6 else valid_rh
                current_cycle["end_rh_avg"] = round(sum(last_6) / len(last_6), 2) if last_6 else current_cycle.get("start_rh", 0)
                currents = current_cycle["_currents"]
                times = current_cycle["_times"]
                energy_ws = 0.0
                for j in range(1, len(currents)):
                    dt = (times[j] - times[j-1]).total_seconds()
                    energy_ws += currents[j-1] * voltage * dt
                current_cycle["energy_kwh"] = round(energy_ws / 3_600_000, 4)
                _confirm_peak()
                current_cycle["current_spike_avg"] = round(sum(_peak_values) / len(_peak_values), 2) if _peak_values else 0.0
                current_cycle["ignition_count"] = len(_peak_values)
                motor_readings = current_cycle.get("_motor_readings", [])
                filter_threshold = None
                if motor_readings:
                    filter_threshold = (sum(motor_readings) / len(motor_readings)) * 1.15
                median_val = _compute_motor_baseline_median(motor_readings, filter_threshold=filter_threshold)
                current_cycle["motor_baseline_median"] = round(median_val, 3)
                del current_cycle["_rh_history"]
                del current_cycle["_currents"]
                del current_cycle["_times"]
                del current_cycle["_motor_readings"]
                current_cycle["start_time"] = current_cycle["start_time"].isoformat()
                current_cycle["end_time"] = current_cycle["end_time"].isoformat()
                cycles.append(current_cycle)

                current_cycle = {}
                _prev_current = 0.0
                _peak_values = []
        if in_cycle:
            in_cycle = False
            last_r = readings[-1]
            current_cycle["end_time"] = last_r[0]
            duration = current_cycle["end_time"] - current_cycle["start_time"]
            current_cycle["duration_minutes"] = round(duration.total_seconds() / 60, 1)
            valid_rh = [v for v in current_cycle["_rh_history"] if v is not None]
            first_6 = valid_rh[:6] if len(valid_rh) > 6 else valid_rh
            current_cycle["start_rh"] = round(sum(first_6) / len(first_6), 2) if first_6 else None
            last_6 = valid_rh[-6:] if len(valid_rh) > 6 else valid_rh
            current_cycle["end_rh_avg"] = round(sum(last_6) / len(last_6), 2) if last_6 else current_cycle.get("start_rh", 0)
            _confirm_peak()
            currents = current_cycle["_currents"]
            times = current_cycle["_times"]
            energy_ws = 0.0
            for j in range(1, len(currents)):
                dt = (times[j] - times[j-1]).total_seconds()
                energy_ws += currents[j-1] * voltage * dt
            current_cycle["energy_kwh"] = round(energy_ws / 3_600_000, 4)
            current_cycle["current_spike_avg"] = round(sum(_peak_values) / len(_peak_values), 2) if _peak_values else 0.0
            current_cycle["ignition_count"] = len(_peak_values)
            motor_readings = current_cycle.get("_motor_readings", [])
            filter_threshold = None
            if motor_readings:
                filter_threshold = (sum(motor_readings) / len(motor_readings)) * 1.15
            median_val = _compute_motor_baseline_median(motor_readings, filter_threshold=filter_threshold)
            current_cycle["motor_baseline_median"] = round(median_val, 3)
            del current_cycle["_rh_history"]
            del current_cycle["_currents"]
            del current_cycle["_times"]
            del current_cycle["_motor_readings"]
            current_cycle["start_time"] = current_cycle["start_time"].isoformat()
            current_cycle["end_time"] = current_cycle["end_time"].isoformat()
            cycles.append(current_cycle)

            current_cycle = {}

        cycles = [c for c in cycles if c.get("duration_minutes", 0) >= 1.0]

        return jsonify(cycles)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cur.close()
            release_conn(conn)

# --- API: MONTHLY ENERGY SUMMARY ---
@app.route('/api/energy_summary')
@login_required
def api_energy_summary():
    """Return monthly energy consumption per appliance for the current user."""
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    try:
        year, month = map(int, month_str.split('-'))
        month_start = datetime(year, month, 1)
        if month == 12:
            month_end = datetime(year + 1, 1, 1)
        else:
            month_end = datetime(year, month + 1, 1)
    except ValueError:
        return jsonify({'error': 'Invalid month format. Use YYYY-MM.'}), 400

    conn = get_conn()
    if not conn:
        return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    try:
        # Get all appliances for current user
        cur.execute("""
            SELECT id, name, type, created_at
            FROM appliances WHERE user_id = %s ORDER BY name
        """, (current_user.id,))
        appliances = cur.fetchall()

        result = []
        by_type = defaultdict(float)
        total_kwh = 0.0

        for app_id, app_name, app_type, created_at in appliances:
            voltage = get_appliance_voltage(app_id)
            # Clamp query to appliance creation time
            query_start = max(month_start, created_at) if created_at else month_start

            if 'Dryer' in app_type:
                cur.execute("""
                    SELECT r.time, r.imotor
                    FROM dryer_readings r
                    JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                    WHERE sn.appliance_id = %s AND r.time >= %s AND r.time < %s
                    ORDER BY r.time ASC
                """, (app_id, query_start, month_end))
            else:
                cur.execute("""
                    SELECT r.time, r.icompressor
                    FROM hvac_readings r
                    JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                    WHERE sn.appliance_id = %s AND r.time >= %s AND r.time < %s
                    ORDER BY r.time ASC
                """, (app_id, query_start, month_end))

            readings = cur.fetchall()
            energy = _compute_energy_kwh(readings, voltage)
            result.append({
                'id': app_id,
                'name': app_name,
                'type': app_type,
                'energy_kwh': energy
            })
            by_type[app_type] += energy
            total_kwh += energy

        return jsonify({
            'month': month_str,
            'appliances': result,
            'total_kwh': round(total_kwh, 4),
            'by_type': dict(by_type)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cur.close()
            release_conn(conn)


@app.route('/api/energy_summary/export')
@login_required
def api_energy_summary_export():
    """Export monthly energy summary to Excel."""
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    try:
        year, month = map(int, month_str.split('-'))
        month_start = datetime(year, month, 1)
        if month == 12:
            month_end = datetime(year + 1, 1, 1)
        else:
            month_end = datetime(year, month + 1, 1)
    except ValueError:
        return jsonify({'error': 'Invalid month format. Use YYYY-MM.'}), 400

    conn = get_conn()
    if not conn:
        return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, name, type, created_at
            FROM appliances WHERE user_id = %s ORDER BY type, name
        """, (current_user.id,))
        appliances = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Energy Summary"
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Title rows
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Monthly Energy Consumption Summary"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Month: {month_str}"
        ws['A2'].font = Font(size=11, bold=True)
        ws.merge_cells('A3:E3')
        ws['A3'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Headers
        headers = ["Appliance Type", "Appliance Name", "Energy Consumption (kWh)"]
        row_idx = 5
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        total_kwh = 0.0
        for app_id, app_name, app_type, created_at in appliances:
            voltage = get_appliance_voltage(app_id)
            query_start = max(month_start, created_at) if created_at else month_start

            if 'Dryer' in app_type:
                cur.execute("""
                    SELECT r.time, r.imotor
                    FROM dryer_readings r
                    JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                    WHERE sn.appliance_id = %s AND r.time >= %s AND r.time < %s
                    ORDER BY r.time ASC
                """, (app_id, query_start, month_end))
            else:
                cur.execute("""
                    SELECT r.time, r.icompressor
                    FROM hvac_readings r
                    JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                    WHERE sn.appliance_id = %s AND r.time >= %s AND r.time < %s
                    ORDER BY r.time ASC
                """, (app_id, query_start, month_end))

            readings = cur.fetchall()
            energy = _compute_energy_kwh(readings, voltage)
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=app_type)
            ws.cell(row=row_idx, column=2, value=app_name)
            ws.cell(row=row_idx, column=3, value=energy)
            total_kwh += energy

        # Total row
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="Total")
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value=round(total_kwh, 4))
        ws.cell(row=row_idx, column=3).font = Font(bold=True)

        # Auto-width
        for col_idx in range(1, 4):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for r in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=r, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Energy_Summary_{month_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cur.close()
            release_conn(conn)


# --- API: ENERGY MONTHS (only months with data) ---
@app.route('/api/energy_months')
@login_required
def api_energy_months():
    """Return distinct year-months where the current user has sensor readings."""
    conn = get_conn()
    if not conn:
        return jsonify([]), 500
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT TO_CHAR(DATE_TRUNC('month', r.time), 'YYYY-MM') AS month
            FROM hvac_readings r
            JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE a.user_id = %s
            UNION
            SELECT DISTINCT TO_CHAR(DATE_TRUNC('month', r.time), 'YYYY-MM') AS month
            FROM dryer_readings r
            JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE a.user_id = %s
            ORDER BY month DESC
        """, (current_user.id, current_user.id))
        rows = cur.fetchall()
        return jsonify([r[0] for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cur.close()
            release_conn(conn)


# --- DISCORD WEBHOOK API ---
@app.route('/api/user/discord_webhook', methods=['GET'])
@login_required
def api_discord_webhook_get():
    conn = get_conn()
    if not conn:
        return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    try:
        cur.execute("SELECT discord_webhook_url FROM users WHERE id = %s", (current_user.id,))
        row = cur.fetchone()
        url = row[0] if row and row[0] else ''
        masked = '...' + url[-20:] if len(url) > 20 else url
        return jsonify({'url': url, 'masked': masked, 'configured': bool(url)})
    finally:
        cur.close()
        release_conn(conn)

@app.route('/api/user/discord_webhook', methods=['POST'])
@login_required
def api_discord_webhook_post():
    data = request.get_json() or {}
    url = data.get('url', '').strip()
    conn = get_conn()
    if not conn:
        return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    try:
        cur.execute("UPDATE users SET discord_webhook_url = %s WHERE id = %s", (url or None, current_user.id))
        conn.commit()
        return jsonify({'success': True, 'configured': bool(url)})
    finally:
        cur.close()
        release_conn(conn)

@app.route('/api/user/discord_webhook/test', methods=['POST'])
@login_required
def api_discord_webhook_test():
    data = request.get_json() or {}
    url = data.get('url', '').strip()
    if not url:
        return jsonify({'error': 'No webhook URL provided'}), 400
    try:
        embed = {
            "title": "✅ Test Alert",
            "description": "Your Discord webhook is configured correctly! Alerts from your IoT monitoring system will appear here.",
            "color": 0x10B981,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "footer": {"text": "IoT Monitoring & Predictive Maintenance"}
        }
        resp = requests.post(url, json={"embeds": [embed]}, timeout=5)
        if resp.status_code in (200, 204):
            return jsonify({'success': True})
        else:
            return jsonify({'error': f'Discord returned status {resp.status_code}'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# --- MQTTv5 persistent session properties ---
try:
    from paho.mqtt.properties import Properties
    from paho.mqtt.packettypes import PacketTypes
    _mqtt_connect_props = Properties(PacketTypes.CONNECT)
    _mqtt_connect_props.SessionExpiryInterval = 3600  # retain session for 1 hour
except Exception:
    _mqtt_connect_props = None

# --- MAIN ---
if __name__ == '__main__':
    mqtt_client.connect(MQTT_HOST, MQTT_PORT, 60, clean_start=False, properties=_mqtt_connect_props)
    mqtt_client.loop_start()
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)

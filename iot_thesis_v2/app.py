import os
import io
import sys
import json
import time
import math
import random
import threading
from datetime import datetime, timedelta, timezone

import psycopg2
import psycopg2.errors
import bcrypt
import paho.mqtt.client as mqtt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from flask import Flask, request, jsonify, render_template, redirect, url_for, flash, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user

app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'iot-thesis-secret-change-this-in-production')

# --- CONFIGURATION ---
MQTT_HOST = os.getenv("MQTT_HOST", "d57bf82836a7485d9b67b270c681fe6e.s1.eu.hivemq.cloud")
MQTT_PORT = int(os.getenv("MQTT_PORT", "8883"))
MQTT_USER = os.getenv("MQTT_USER", "esp32user")
MQTT_PASS = os.getenv("MQTT_PASS", "IoTTHESIS1")

UNPAIRED_CACHE = {}
DEDUPE_CACHE = {}
EVENT_DEDUPE_CACHE = {}
CALIBRATION_TRACKER = {}  # appliance_id -> {start_tcoil, start_time}
CYCLE_TRACKER = {}        # appliance_id -> {start_time, last_time}
SPC_ALERT_COOLDOWN = {}   # (appliance_id, metric_name) -> last_alert_timestamp

CLIENT_ID = f"FlaskBackend_{random.randint(10000, 99999)}"

# --- MQTT Setup ---
mqtt_client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2, client_id=CLIENT_ID, protocol=mqtt.MQTTv5)
mqtt_client.username_pw_set(MQTT_USER, MQTT_PASS)
mqtt_client.tls_set()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

from psycopg2 import pool
try:
    db_pool = psycopg2.pool.SimpleConnectionPool(
        1, 10,
        host=os.getenv("DB_HOST", "localhost"),
        port=int(os.getenv("DB_PORT", "5432")),
        dbname=os.getenv("DB_NAME", "iot_db"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", "IOTTHESIS")
    )
except Exception as e:
    print(f"DB Pool Error: {e}")
    db_pool = None

def get_conn():
    try:
        if db_pool:
            return db_pool.getconn()
    except Exception as e:
        print(f"DB Get Connection Error: {e}")
    return None

def release_conn(conn):
    if db_pool and conn:
        db_pool.putconn(conn)

class User(UserMixin):
    def __init__(self, id, email, name):
        self.id = id
        self.email = email
        self.name = name

@login_manager.user_loader
def load_user(user_id):
    conn = get_conn()
    if not conn: return None
    cur = conn.cursor()
    cur.execute("SELECT id, email, name FROM users WHERE id = %s", (user_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    return User(row[0], row[1], row[2]) if row else None

# --- CALIBRATION HELPERS ---
def get_appliance_calibration(appliance_id):
    conn = get_conn()
    default_cal = {
        'type': 'HVAC',
        't1_m': 1.0, 't1_c': 0.0,
        'h1_m': 1.0, 'h1_c': 0.0,
        't2_m': 1.0, 't2_c': 0.0,
        'h2_m': 1.0, 'h2_c': 0.0,
        'tcoil_c': 0.0, 'icompressor_offset': 0.0, 'tcoil_m': 1.0,
    }
    if not conn: return default_cal
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT type, treturn_slope, treturn_intercept,
                   rhreturn_slope, rhreturn_intercept,
                   tsupply_slope, tsupply_intercept,
                   rhsupply_slope, rhsupply_intercept,
                   tcoil_offset, icompressor_offset, tcoil_slope
            FROM appliances WHERE id = %s
        """, (appliance_id,))
        row = cur.fetchone()
        cur.close()
        release_conn(conn)
        if not row: return default_cal
        return {
            'type': row[0],
            't1_m': row[1] if row[1] is not None else 1.0,
            't1_c': row[2] if row[2] is not None else 0.0,
            'h1_m': row[3] if row[3] is not None else 1.0,
            'h1_c': row[4] if row[4] is not None else 0.0,
            't2_m': row[5] if row[5] is not None else 1.0,
            't2_c': row[6] if row[6] is not None else 0.0,
            'h2_m': row[7] if row[7] is not None else 1.0,
            'h2_c': row[8] if row[8] is not None else 0.0,
            'tcoil_c': row[9] if row[9] is not None else 0.0,
            'icompressor_offset': row[10] if row[10] is not None else 0.0,
            'tcoil_m': row[11] if row[11] is not None else 1.0,
        }
    except Exception as e:
        print(f"Error getting calibration: {e}")
        if conn: release_conn(conn)
        return default_cal

def apply_calibration(raw_val, m, c):
    if raw_val is None: return 0.0
    return (float(raw_val) * float(m)) + float(c)

# --- DATA QUERY HELPERS ---
def latest_row_for_appliance(appliance_id):
    conn = get_conn()
    if not conn: return None, None
    cur = conn.cursor()
    cur.execute("SELECT type, created_at FROM appliances WHERE id = %s", (appliance_id,))
    type_row = cur.fetchone()
    if not type_row:
        cur.close()
        release_conn(conn)
        return None, None
    dev_type, created_at = type_row[0], type_row[1]
    if "Dryer" in dev_type:
        cur.execute("""
            SELECT dr.time, dr.texhaust, dr.rh_exhaust, dr.pressure, dr.imotor
            FROM dryer_readings dr
            JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND dr.time >= a.created_at
            ORDER BY dr.time DESC LIMIT 1
        """, (appliance_id,))
    else:
        cur.execute("""
            SELECT sr.time, sr.treturn, sr.rhreturn, sr.tsupply, sr.rhsupply, sr.tcoil, sr.icompressor
            FROM hvac_readings sr
            JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND sr.time >= a.created_at
            ORDER BY sr.time DESC LIMIT 1
        """, (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    return row, dev_type

def latest_n_rows_for_appliance(appliance_id, limit, start=None, end=None):
    conn = get_conn()
    if not conn: return [], "Unknown"
    cur = conn.cursor()
    cur.execute("SELECT type, created_at FROM appliances WHERE id = %s", (appliance_id,))
    type_row = cur.fetchone()
    if not type_row: return [], "Unknown"
    dev_type, created_at = type_row[0], type_row[1]
    if start and end:
        if "Dryer" in dev_type:
            cur.execute("""
                SELECT dr.time, dr.texhaust, dr.rh_exhaust, dr.pressure, dr.imotor
                FROM dryer_readings dr
                JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND dr.time >= %s AND dr.time <= %s AND dr.time >= %s
                ORDER BY dr.time ASC LIMIT 5000
            """, (appliance_id, start, end, created_at))
        else:
            cur.execute("""
                SELECT sr.time, sr.treturn, sr.rhreturn, sr.tsupply, sr.rhsupply, sr.tcoil, sr.icompressor
                FROM hvac_readings sr
                JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND sr.time >= %s AND sr.time <= %s AND sr.time >= %s
                ORDER BY sr.time ASC LIMIT 5000
            """, (appliance_id, start, end, created_at))
    else:
        if "Dryer" in dev_type:
            cur.execute("""
                SELECT dr.time, dr.texhaust, dr.rh_exhaust, dr.pressure, dr.imotor
                FROM dryer_readings dr
                JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND dr.time >= %s
                ORDER BY dr.time DESC LIMIT %s
            """, (appliance_id, created_at, limit))
        else:
            cur.execute("""
                SELECT sr.time, sr.treturn, sr.rhreturn, sr.tsupply, sr.rhsupply, sr.tcoil, sr.icompressor
                FROM hvac_readings sr
                JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                WHERE sn.appliance_id = %s AND sr.time >= %s
                ORDER BY sr.time DESC LIMIT %s
            """, (appliance_id, created_at, limit))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return (rows[::-1], dev_type) if not (start and end) else (rows, dev_type)

def get_latest_current_for_appliance(appliance_id):
    conn = get_conn()
    if not conn: return 0.0
    cur = conn.cursor()
    try:
        cur.execute("SELECT type, created_at FROM appliances WHERE id = %s", (appliance_id,))
        type_row = cur.fetchone()
        if not type_row: return 0.0
        dev_type, created_at = type_row[0], type_row[1]
        if "Dryer" in dev_type:
            cur.execute("""
                SELECT dr.imotor FROM dryer_readings dr
                JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND dr.time >= a.created_at
                ORDER BY dr.time DESC LIMIT 1
            """, (appliance_id,))
        else:
            cur.execute("""
                SELECT sr.icompressor FROM hvac_readings sr
                JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND sr.time >= a.created_at
                ORDER BY sr.time DESC LIMIT 1
            """, (appliance_id,))
        row = cur.fetchone()
        return float(row[0]) if row and row[0] is not None else 0.0
    except Exception:
        return 0.0
    finally:
        cur.close()
        release_conn(conn)

def get_appliances_for_user(user_id):
    conn = get_conn()
    if not conn: return []
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, type, brand, location, created_at, operational_status, sub_type, baseline_configured
        FROM appliances WHERE user_id = %s ORDER BY created_at
    """, (user_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return [{'id':r[0],'name':r[1],'type':r[2],'brand':r[3],'location':r[4],
             'created_at':r[5], 'status':r[6], 'sub_type':r[7], 'baseline_configured':r[8]} for r in rows]

def unpaired_nodes():
    conn = get_conn()
    if not conn: return []
    cur = conn.cursor()
    cur.execute("""
        SELECT sn.id, sn.mac_address, sn.status
        FROM sensor_nodes sn
        WHERE sn.status = 'unpaired' AND sn.last_seen >= NOW() - INTERVAL '30 seconds'
        ORDER BY sn.id
    """)
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return [{'id':r[0], 'mac_address':r[1], 'status':r[2], 'readings':{}} for r in rows]

def get_all_nodes_for_user(user_id):
    conn = get_conn()
    if not conn: return []
    cur = conn.cursor()
    cur.execute("""
        SELECT sn.id, sn.mac_address, sn.status, a.name
        FROM sensor_nodes sn
        LEFT JOIN appliances a ON sn.appliance_id = a.id
        WHERE a.user_id = %s OR sn.status = 'unpaired'
        ORDER BY sn.id
    """, (user_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return [{'id':r[0], 'mac_address':r[1], 'status':r[2], 'appliance':r[3] or 'Unpaired'} for r in rows]

def send_node_command(mac, command_str):
    if not mqtt_client.is_connected():
        print(f"Cannot send command {command_str} to {mac}, MQTT disconnected.")
        return
    mqtt_client.publish(f"iot/nodes/{mac}/control", command_str)
    print(f"Backend -> Node {mac}: {command_str}")

# --- SPC BASELINE HELPERS ---
HVAC_METRICS = ['deltat', 'deltarh', 'tcoil', 'rhreturn', 'rhsupply', 'current']
DRYER_METRICS = ['texhaust', 'rhexhaust', 'pressure', 'current']

def get_spc_baselines(appliance_id):
    """Fetch manual SPC baselines from spc_manual_baselines table."""
    conn = get_conn()
    if not conn: return {}
    cur = conn.cursor()
    cur.execute("""
        SELECT metric_name, ucl, lcl, mean FROM spc_manual_baselines WHERE appliance_id = %s
    """, (appliance_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    result = {}
    for r in rows:
        result[r[0]] = {'ucl': float(r[1]), 'lcl': float(r[2]), 'mean': float(r[3])}
    return result

def save_spc_baselines(appliance_id, baselines):
    """baselines: dict of metric_name -> {ucl, lcl}"""
    conn = get_conn()
    if not conn: return False, "DB connection error"
    cur = conn.cursor()
    try:
        for metric, vals in baselines.items():
            ucl = float(vals['ucl'])
            lcl = float(vals['lcl'])
            if ucl <= lcl:
                return False, f"UCL must be greater than LCL for {metric}"
            mean = (ucl + lcl) / 2.0
            cur.execute("""
                INSERT INTO spc_manual_baselines (appliance_id, metric_name, ucl, lcl, mean, updated_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
                ON CONFLICT (appliance_id, metric_name) DO UPDATE SET
                    ucl = EXCLUDED.ucl,
                    lcl = EXCLUDED.lcl,
                    mean = EXCLUDED.mean,
                    updated_at = NOW()
            """, (appliance_id, metric, ucl, lcl, mean))
        cur.execute("UPDATE appliances SET baseline_configured = TRUE WHERE id = %s", (appliance_id,))
        conn.commit()
        return True, "Baseline saved successfully"
    except Exception as e:
        conn.rollback()
        return False, str(e)
    finally:
        cur.close()
        release_conn(conn)

def notify_node_baseline_set(appliance_id):
    conn = get_conn()
    if not conn: return
    cur = conn.cursor()
    cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s", (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if row and row[0]:
        send_node_command(row[0], "baseline:set")

def check_spc_alerts(appliance_id, reading_data, dev_type):
    """Check running telemetry against SPC limits and insert alerts if breached."""
    baselines = get_spc_baselines(appliance_id)
    if not baselines:
        return
    conn = get_conn()
    if not conn: return
    cur = conn.cursor()
    try:
        cur.execute("SELECT alert_enabled FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        if not row or not row[0]:
            return
        alert_enabled = row[0]
        if not alert_enabled:
            return

        now = datetime.now(timezone.utc)
        metrics_to_check = []

        if "Dryer" in dev_type:
            metrics_to_check = [
                ('texhaust', reading_data.get('texhaust')),
                ('rhexhaust', reading_data.get('rhexhaust')),
                ('pressure', reading_data.get('pressure')),
                ('current', reading_data.get('current')),
            ]
        else:
            metrics_to_check = [
                ('deltat', reading_data.get('deltat')),
                ('deltarh', reading_data.get('deltarh')),
                ('tcoil', reading_data.get('tcoil')),
                ('rhreturn', reading_data.get('rhreturn')),
                ('rhsupply', reading_data.get('rhsupply')),
                ('current', reading_data.get('current')),
            ]

        for metric_name, value in metrics_to_check:
            if value is None or metric_name not in baselines:
                continue
            bl = baselines[metric_name]
            ucl = bl['ucl']
            lcl = bl['lcl']
            val = float(value)

            # Rate limit: one alert per metric per 5 minutes
            cooldown_key = (appliance_id, metric_name)
            if cooldown_key in SPC_ALERT_COOLDOWN:
                if (now - SPC_ALERT_COOLDOWN[cooldown_key]).total_seconds() < 300:
                    continue

            alert_type = None
            threshold = None
            message = None

            if val > ucl:
                alert_type = 'spc_ucl_breach'
                threshold = ucl
                message = f"{metric_name.upper()} {val:.2f} exceeds UCL {ucl:.2f}"
            elif val < lcl:
                alert_type = 'spc_lcl_breach'
                threshold = lcl
                message = f"{metric_name.upper()} {val:.2f} below LCL {lcl:.2f}"

            if alert_type:
                cur.execute("""
                    INSERT INTO alerts (appliance_id, alert_type, message, value, threshold, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (appliance_id, alert_type, message, val, threshold, now))
                conn.commit()
                SPC_ALERT_COOLDOWN[cooldown_key] = now
    except Exception as e:
        print(f"SPC alert check error: {e}")
    finally:
        if conn:
            cur.close()
            release_conn(conn)

# --- MQTT HANDLERS ---
def on_mqtt_connect(client, userdata, flags, rc, properties=None):
    if rc == 0:
        print(f"MQTT Connected Successfully! ID: {client._client_id}")
        client.subscribe("iot/nodes/+/events")
        client.subscribe("iot/nodes/+/telemetry")
    else:
        print(f"MQTT Connection Failed with Code {rc}")

def handle_node_events(mac, payload):
    try:
        data = json.loads(payload)
        event_type = data.get("event")
        event_hash = f"{mac}_{event_type}"
        now_time = time.time()
        if event_hash in EVENT_DEDUPE_CACHE:
            if now_time - EVENT_DEDUPE_CACHE[event_hash] < 5.0:
                print(f"Duplicate event {event_type} from {mac} ignored (in-memory dedup).")
                return
        EVENT_DEDUPE_CACHE[event_hash] = now_time

        conn = get_conn()
        if not conn: return
        cur = conn.cursor()
        cur.execute("SELECT appliance_id, status FROM sensor_nodes WHERE mac_address = %s", (mac,))
        node_row = cur.fetchone()

        if event_type == "event_request_config":
            if not node_row:
                cur.execute(
                    "INSERT INTO sensor_nodes (mac_address, status, created_at, last_seen) VALUES (%s, 'unpaired', NOW(), NOW())",
                    (mac,)
                )
                conn.commit()
                print(f"Auto-registered unknown node {mac}")
            else:
                cur.execute(
                    "UPDATE sensor_nodes SET last_seen = NOW() WHERE mac_address = %s",
                    (mac,)
                )
                conn.commit()
            if not node_row or not node_row[0] or (node_row and node_row[1] != 'paired'):
                send_node_command(mac, "settype:unpaired")
            else:
                appliance_id = node_row[0]
                cur.execute("SELECT type, operational_status FROM appliances WHERE id = %s", (appliance_id,))
                row = cur.fetchone()
                if row:
                    app_type, app_status = row[0], row[1]
                    if app_status == 'calibrating':
                        cur.execute("UPDATE appliances SET operational_status='calibration_needed' WHERE id=%s", (appliance_id,))
                        conn.commit()
                        app_status = 'calibration_needed'
                    if "Dryer" in app_type:
                        send_node_command(mac, "settype:dryer")
                        send_node_command(mac, "restore:normal")
                    else:
                        send_node_command(mac, "settype:hvac")
                        if app_status == 'normal':
                            send_node_command(mac, "restore:normal")
                        else:
                            send_node_command(mac, "restore:calibrationneeded")
            cur.close()
            release_conn(conn)
            return

        if event_type == "checkin":
            cur.execute("UPDATE sensor_nodes SET last_seen = NOW() WHERE mac_address = %s", (mac,))
            conn.commit()
            cur.close()
            release_conn(conn)
            return

        if not node_row or not node_row[0] or node_row[1] != 'paired':
            if event_type in ["event_button2_action_request"]:
                send_node_command(mac, "actiondenied:busy")
            else:
                send_node_command(mac, "baselinefailack")
            cur.close()
            release_conn(conn)
            return

        appliance_id = node_row[0]
        cur.execute("SELECT operational_status, type FROM appliances WHERE id = %s", (appliance_id,))
        stat_row = cur.fetchone()
        if not stat_row:
            cur.close()
            release_conn(conn)
            return
        status, app_type = stat_row[0], stat_row[1]

        if event_type == "event_button2_calibration_request":
            if "Dryer" in app_type:
                send_node_command(mac, "calibrationfailack")
                print(f"Node {mac} (Dryer) Calibration not supported.")
            else:
                if status == 'calibration_needed':
                    cur.execute("UPDATE appliances SET operational_status = 'calibrating', calibration_started_at = NOW() WHERE id = %s", (appliance_id,))
                    conn.commit()
                    cur.execute("""
                        SELECT sr.tcoil FROM hvac_readings sr
                        JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
                        WHERE sn.appliance_id = %s ORDER BY sr.time DESC LIMIT 1
                    """, (appliance_id,))
                    tcoil_row = cur.fetchone()
                    start_tcoil = float(tcoil_row[0]) if tcoil_row and tcoil_row[0] is not None else 25.0
                    CALIBRATION_TRACKER[appliance_id] = {'start_tcoil': start_tcoil, 'start_time': datetime.now()}
                    send_node_command(mac, "startcalibration")
                    print(f"Node {mac} (HVAC) Calibration STARTED. Tcoil start: {start_tcoil}")
                elif status == 'calibrating':
                    send_node_command(mac, "actiondenied:busy")
                    print(f"Node {mac} Action denied. Device is currently busy.")
                else:
                    send_node_command(mac, "calibrationfailack")
                    print(f"Node {mac} Calibration already done or not needed.")

        elif event_type == "calibration_success_request":
            if status == 'calibrating':
                if appliance_id in CALIBRATION_TRACKER:
                    del CALIBRATION_TRACKER[appliance_id]
                base_data = data.get("base")
                final_data = data.get("final")
                if base_data and final_data:
                    try:
                        t1_delta = abs(float(base_data.get("t1", 0)) - float(final_data.get("t1", 0)))
                        t2_delta = abs(float(base_data.get("t2", 0)) - float(final_data.get("t2", 0)))
                        t3_delta = abs(float(base_data.get("t3", 0)) - float(final_data.get("t3", 0)))
                    except (ValueError, TypeError):
                        t1_delta = t2_delta = t3_delta = 0.0
                    if t3_delta < 7.5 or t2_delta < 2.5 or t1_delta < 2.5:
                        cur.execute("UPDATE appliances SET operational_status='calibration_needed' WHERE id=%s", (appliance_id,))
                        conn.commit()
                        send_node_command(mac, "calibrationfailack")
                        print(f"Node {mac} Calibration REJECTED. Sensors didn't drop enough.")
                    else:
                        def safe_polyfit(x1, x2, y1, y2):
                            try:
                                x1, x2, y1, y2 = float(x1), float(x2), float(y1), float(y2)
                                if abs(x1 - x2) < 0.1: return 1.0, 0.0
                                coeffs = np.polyfit([x1, x2], [y1, y2], 1)
                                return float(coeffs[0]), float(coeffs[1])
                            except Exception:
                                return 1.0, 0.0
                        t1_m, t1_c = safe_polyfit(base_data.get("t1", 0), final_data.get("t1", 0), base_data.get("t3", 0), final_data.get("t3", 0))
                        t2_m, t2_c = safe_polyfit(base_data.get("t2", 0), final_data.get("t2", 0), base_data.get("t3", 0), final_data.get("t3", 0))
                        h2_m, h2_c = safe_polyfit(base_data.get("h2", 0), final_data.get("h2", 0), base_data.get("h1", 0), final_data.get("h1", 0))
                        cur.execute("""
                            UPDATE appliances SET
                                treturn_slope=%s, treturn_intercept=%s,
                                rhreturn_slope=1.0, rhreturn_intercept=0.0,
                                tsupply_slope=%s, tsupply_intercept=%s,
                                rhsupply_slope=%s, rhsupply_intercept=%s,
                                tcoil_slope=1.0, tcoil_offset=0.0,
                                operational_status='normal'
                            WHERE id=%s AND operational_status='calibrating'
                        """, (t1_m, t1_c, t2_m, t2_c, h2_m, h2_c, appliance_id))
                        conn.commit()
                        send_node_command(mac, "calibrationsuccessack")
                        print(f"Node {mac} Multi-Point Linear Calibration SUCCESS.")
                else:
                    send_node_command(mac, "calibrationfailack")

        elif event_type == "calibration_fail_request":
            if status == 'calibrating':
                if appliance_id in CALIBRATION_TRACKER:
                    del CALIBRATION_TRACKER[appliance_id]
                cur.execute("UPDATE appliances SET operational_status = 'calibration_needed' WHERE id = %s", (appliance_id,))
                conn.commit()
                send_node_command(mac, "calibrationfailack")

        elif event_type == "maintenance_request":
            if "Dryer" in app_type or status == 'normal':
                cur.execute("INSERT INTO sensor_events (sensor_node_mac, event_type, timestamp) VALUES (%s,%s,%s)",
                            (mac, 'maintenance', datetime.now()))
                conn.commit()
                send_node_command(mac, "maintenanceack")
                print(f"Node {mac} maintenance logged.")
            else:
                send_node_command(mac, "maintenancedenied")
                print(f"Node {mac} maintenance denied: status={status}")

        cur.close()
        release_conn(conn)
    except Exception as e:
        print(f"Error handling event: {e}")

def on_mqtt_message(client, userdata, msg):
    try:
        payload = msg.payload.decode()
        topic = msg.topic
        parts = topic.split("/")
        if len(parts) < 4: return
        mac = parts[2]
        if "events" in topic:
            handle_node_events(mac, payload)
            return
        elif "telemetry" in topic:
            safe_str = payload.replace('nan', 'null')
            data = json.loads(safe_str)
            now_utc = datetime.now(timezone.utc)
            if "agoms" in data:
                actual_time = now_utc - timedelta(milliseconds=max(0, int(data["agoms"])))
            elif "ago_ms" in data:
                actual_time = now_utc - timedelta(milliseconds=max(0, int(data["ago_ms"])))
            else:
                actual_time = now_utc - timedelta(seconds=max(0, int(data.get("ago", 0))))
            if actual_time > now_utc + timedelta(minutes=1):
                actual_time = now_utc
            if mac in DEDUPE_CACHE:
                diff = abs((actual_time - DEDUPE_CACHE[mac]).total_seconds())
                if diff < 1.0:
                    return
            DEDUPE_CACHE[mac] = actual_time

            conn = get_conn()
            if conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT sn.id, sn.appliance_id, a.type, a.operational_status
                    FROM sensor_nodes sn
                    LEFT JOIN appliances a ON sn.appliance_id = a.id
                    WHERE sn.mac_address = %s
                """, (mac,))
                row = cur.fetchone()
                if not row:
                    cur.execute("INSERT INTO sensor_nodes (mac_address, status, last_seen) VALUES (%s, 'unpaired', NOW()) RETURNING id", (mac,))
                    sensor_node_id = cur.fetchone()[0]
                    appliance_id = None
                    appliance_type = None
                    conn.commit()
                else:
                    sensor_node_id, appliance_id, appliance_type, _ = row
                    cur.execute("UPDATE sensor_nodes SET last_seen = NOW() WHERE id = %s", (sensor_node_id,))
                    conn.commit()

                raw_amps = data.get("CurrentA", 0.0)
                final_amps = 0.0
                if raw_amps is not None:
                    try:
                        final_amps = float(raw_amps)
                    except:
                        final_amps = 0.0

                status_field = data.get("status", "running")
                is_running = (status_field == "running")

                reading_values = {}
                if appliance_type:
                    if is_running:
                        if "Dryer" in appliance_type:
                            tex = data.get("BME280Temp")
                            rhex = data.get("BME280Hum")
                            pres = data.get("BME280Pres")
                            cur.execute("""
                                INSERT INTO dryer_readings (sensor_node_id, texhaust, rh_exhaust, pressure, imotor, time)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, (sensor_node_id, tex, rhex, pres, final_amps, actual_time))
                            reading_values = {'texhaust': tex, 'rhexhaust': rhex, 'pressure': pres, 'current': final_amps}
                        else:
                            t1 = data.get("DHT1Temp")
                            h1 = data.get("DHT1Hum")
                            t2 = data.get("DHT2Temp")
                            h2 = data.get("DHT2Hum")
                            t3 = data.get("DS18B20Temp")
                            cur.execute("""
                                INSERT INTO hvac_readings (sensor_node_id, treturn, rhreturn, tsupply, rhsupply, tcoil, icompressor, time)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """, (sensor_node_id, t1, h1, t2, h2, t3, final_amps, actual_time))
                            # Compute deltas for SPC alert checking
                            cal = get_appliance_calibration(appliance_id)
                            t1c = apply_calibration(t1, cal['t1_m'], cal['t1_c'])
                            t2c = apply_calibration(t2, cal['t2_m'], cal['t2_c'])
                            h1c = apply_calibration(h1, cal['h1_m'], cal['h1_c'])
                            h2c = apply_calibration(h2, cal['h2_m'], cal['h2_c'])
                            reading_values = {
                                'deltat': abs(t1c - t2c),
                                'deltarh': abs(h1c - h2c),
                                'tcoil': apply_calibration(t3, cal['tcoil_m'], cal['tcoil_c']),
                                'rhreturn': h1c,
                                'rhsupply': h2c,
                                'current': final_amps
                            }
                        cur.execute("UPDATE sensor_nodes SET last_seen = NOW() WHERE id = %s", (sensor_node_id,))
                        conn.commit()

                        # --- NEW: Real-time SPC alert checking ---
                        if appliance_id and reading_values:
                            check_spc_alerts(appliance_id, reading_values, appliance_type)
                    else:
                        cur.execute("UPDATE sensor_nodes SET last_seen = NOW() WHERE id = %s", (sensor_node_id,))
                        conn.commit()

                    # --- Dryer cycle tracker + end-of-cycle humidity alert ---
                    def _process_cycle_end(app_id, cyc_start, cyc_end):
                        cur.execute("SELECT baseline_configured, alert_enabled FROM appliances WHERE id = %s", (app_id,))
                        app_info = cur.fetchone()
                        if not app_info:
                            return
                        baseline_configured, alert_enabled = app_info[0], app_info[1]
                        if not baseline_configured or not alert_enabled:
                            return
                        cur.execute("""
                            SELECT AVG(rh_exhaust) FROM dryer_readings dr
                            JOIN sensor_nodes sn ON dr.sensor_node_id = sn.id
                            WHERE sn.appliance_id = %s AND dr.time >= %s AND dr.time <= %s
                        """, (app_id, cyc_end - timedelta(minutes=2), cyc_end))
                        avg_rh_row = cur.fetchone()
                        avg_rh = float(avg_rh_row[0]) if avg_rh_row and avg_rh_row[0] is not None else None
                        cur.execute("SELECT alert_rhexhaust_threshold FROM appliances WHERE id = %s", (app_id,))
                        thr_row = cur.fetchone()
                        threshold = float(thr_row[0]) if thr_row and thr_row[0] is not None else 40.0
                        if avg_rh is not None and avg_rh > threshold:
                            cur.execute("""
                                INSERT INTO alerts (appliance_id, alert_type, message, value, threshold, cycle_start_time, cycle_end_time)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """, (app_id, 'dryer_humidity_high',
                                  f"Dry cycle incomplete — exhaust humidity {avg_rh:.1f}% exceeds threshold {threshold:.1f}%",
                                  avg_rh, threshold, cyc_start, cyc_end))
                            conn.commit()

                    if "Dryer" in appliance_type:
                        if is_running:
                            if appliance_id in CYCLE_TRACKER:
                                tracker = CYCLE_TRACKER[appliance_id]
                                if (actual_time - tracker["last_time"]).total_seconds() > 60:
                                    _process_cycle_end(appliance_id, tracker.get("start_time", tracker["last_time"]), tracker["last_time"])
                                    del CYCLE_TRACKER[appliance_id]
                            tracker = CYCLE_TRACKER.setdefault(appliance_id, {"start_time": actual_time, "last_time": actual_time})
                            tracker["last_time"] = actual_time
                        else:
                            if appliance_id in CYCLE_TRACKER:
                                tracker = CYCLE_TRACKER[appliance_id]
                                _process_cycle_end(appliance_id, tracker.get("start_time", tracker["last_time"]), tracker["last_time"])
                                del CYCLE_TRACKER[appliance_id]
                else:
                    UNPAIRED_CACHE[sensor_node_id] = {
                        "data": data,
                        "amps": final_amps,
                        "time": actual_time,
                        "mac": mac
                    }
                cur.close()
                release_conn(conn)
    except Exception as e:
        print(f"MQTT Message Error: {e}")

mqtt_client.on_connect = on_mqtt_connect
mqtt_client.on_message = on_mqtt_message

# --- AUTH ROUTES ---
@app.route('/')
def landing():
    return redirect(url_for('dashboard')) if current_user.is_authenticated else redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html',
                           user=current_user,
                           appliances=get_appliances_for_user(current_user.id),
                           unpaired_nodes=unpaired_nodes(),
                           all_nodes=get_all_nodes_for_user(current_user.id))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'GET':
        return render_template('signup.html')
    name = request.form.get('name')
    email = request.form.get('email')
    password = request.form.get('password')
    if not name or not email or not password:
        flash('All fields are required', 'error')
        return redirect(url_for('signup'))
    hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    conn = get_conn()
    if conn:
        cur = conn.cursor()
        try:
            cur.execute("INSERT INTO users (email, password_hash, name) VALUES (%s, %s, %s) RETURNING id", (email, hashed, name))
            user_id = cur.fetchone()[0]
            conn.commit()
            user = User(user_id, email, name)
            login_user(user)
            return redirect(url_for('dashboard'))
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
            flash('Email already exists', 'error')
        except Exception as e:
            conn.rollback()
            flash(f'Error: {e}', 'error')
        finally:
            cur.close()
            release_conn(conn)
    return redirect(url_for('signup'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')
    email = request.form.get('email')
    password = request.form.get('password')
    conn = get_conn()
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT id, email, name, password_hash FROM users WHERE email = %s", (email,))
        row = cur.fetchone()
        cur.close()
        release_conn(conn)
        if row and bcrypt.checkpw(password.encode('utf-8'), row[3].encode('utf-8')):
            user = User(row[0], row[1], row[2])
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Invalid credentials', 'error')
    return redirect(url_for('login'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('landing'))

# --- PAIRING ROUTES ---
@app.route('/devices/pair', methods=['POST'])
@login_required
def pair_device():
    name = request.form.get('name')
    dev_type = request.form.get('type')
    node_id = request.form.get('node_id')
    sub_type = request.form.get('sub_type', 'noninverter')
    if not name or not dev_type or not node_id:
        flash('All fields are required', 'error')
        return redirect(url_for('dashboard'))
    conn = get_conn()
    cur = conn.cursor()
    try:
        initial_status = 'normal' if "Dryer" in dev_type else 'calibration_needed'
        default_offset = -0.111 if "Dryer" in dev_type else -0.033
        cur.execute("""
            INSERT INTO appliances (user_id, name, type, location, brand, operational_status, sub_type, icompressor_offset)
            VALUES (%s, %s, %s, 'Home', 'Generic', %s, %s, %s) RETURNING id
        """, (current_user.id, name, dev_type, initial_status, sub_type, default_offset))
        appliance_id = cur.fetchone()[0]
        cur.execute("UPDATE sensor_nodes SET appliance_id = %s, status = 'paired' WHERE id = %s RETURNING mac_address", (appliance_id, node_id))
        mac = cur.fetchone()[0]
        if int(node_id) in UNPAIRED_CACHE:
            del UNPAIRED_CACHE[int(node_id)]
        conn.commit()
        cmd = "settype:dryer" if "Dryer" in dev_type else "settype:hvac"
        send_node_command(mac, cmd)
        if "Dryer" in dev_type:
            send_node_command(mac, "restore:normal")
            flash(f'{name} added! Device is ready. Configure baseline to enable alerts.', 'success')
        else:
            send_node_command(mac, "restore:calibrationneeded")
            flash(f'{name} added! Calibrate sensors, then configure baseline.', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error pairing device: {e}', 'error')
    finally:
        cur.close()
        release_conn(conn)
    return redirect(url_for('dashboard'))

@app.route('/devices/<int:appliance_id>/forget', methods=['POST'])
@login_required
def forget_device(appliance_id):
    conn = get_conn()
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s", (appliance_id,))
        mac_row = cur.fetchone()
        cur.execute("UPDATE sensor_nodes SET appliance_id = NULL, status = 'unpaired' WHERE appliance_id = %s", (appliance_id,))
        cur.execute("DELETE FROM appliances WHERE id = %s", (appliance_id,))
        conn.commit()
        cur.close()
        release_conn(conn)
        if mac_row:
            send_node_command(mac_row[0], "settype:unpaired")
            print(f"Device forgotten: sent settype:unpaired to {mac_row[0]}")
        flash('Device forgotten.', 'success')
    return redirect(url_for('dashboard'))

# --- API: UNPAIRED NODES ---
@app.route('/api/unpaired_nodes')
@login_required
def api_unpaired():
    return jsonify(unpaired_nodes())

@app.route('/api/node/<int:node_id>/latest')
@login_required
def api_node_latest(node_id):
    if node_id in UNPAIRED_CACHE:
        cache = UNPAIRED_CACHE[node_id]
        d = cache["data"]
        if "BME280Temp" in d:
            return jsonify({
                "time": cache["time"].isoformat(),
                "Texhaust": d.get("BME280Temp", 0),
                "RHexhaust": d.get("BME280Hum", 0),
                "Pressure": d.get("BME280Pres", 0),
                "Imotor": cache.get("amps", 0),
                "cal_state": d.get("cal_state", "idle")
            })
        else:
            return jsonify({
                "time": cache["time"].isoformat(),
                "Treturn": d.get("DHT1Temp", 0),
                "RHreturn": d.get("DHT1Hum", 0),
                "Tsupply": d.get("DHT2Temp", 0),
                "RHsupply": d.get("DHT2Hum", 0),
                "Tcoil": d.get("DS18B20Temp", 0),
                "Icompressor": cache.get("amps", 0),
                "cal_state": d.get("cal_state", "idle")
            })
    return jsonify({'error': 'Node not in cache'}), 404

# --- API: DEVICE DATA ---
@app.route('/api/device/<int:appliance_id>/latest')
@login_required
def api_device_latest(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("""
        SELECT a.operational_status, a.type, sn.last_seen, a.baseline_configured
        FROM appliances a
        LEFT JOIN sensor_nodes sn ON sn.appliance_id = a.id
        WHERE a.id = %s
    """, (appliance_id,))
    status_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not status_row: return jsonify({'error': 'not found'}), 404

    operational_status = status_row[0]
    dev_type = status_row[1]
    last_seen = status_row[2]
    baseline_configured = status_row[3]
    is_calibrated = operational_status not in ['calibration_needed', 'calibrating']

    offline_threshold_seconds = 600
    is_offline = False
    now = datetime.now(timezone.utc)
    if last_seen:
        if last_seen.tzinfo is None:
            last_seen = last_seen.replace(tzinfo=timezone.utc)
        is_offline = (now - last_seen).total_seconds() > offline_threshold_seconds
    else:
        is_offline = True

    if "Dryer" not in dev_type and operational_status in ['calibration_needed', 'calibrating']:
        return jsonify({
            'error': 'Sensor not calibrated',
            'calibrated': False,
            'status': operational_status,
            'type': dev_type,
            'running_status': 'idle',
            'is_offline': is_offline,
            'has_data': False,
            'baseline_configured': baseline_configured
        }), 200

    row_data, dev_type = latest_row_for_appliance(appliance_id)
    cal = get_appliance_calibration(appliance_id)

    if not row_data or 't1_m' not in cal:
        return jsonify({
            'type': dev_type,
            'calibrated': is_calibrated and ('t1_m' in cal),
            'status': operational_status,
            'running_status': 'idle',
            'is_offline': is_offline,
            'has_data': False,
            'baseline_configured': baseline_configured
        }), 200

    time_val = row_data[0]
    stale_threshold_seconds = 60
    if time_val.tzinfo is None:
        time_val = time_val.replace(tzinfo=timezone.utc)
    is_stale = time_val > now or (now - time_val).total_seconds() > stale_threshold_seconds

    if "Dryer" in dev_type:
        imotor = max(0.0, row_data[4] or 0)
        running_status = 'idle' if is_stale else ('running' if imotor >= 0.4 else 'idle')
        return jsonify({
            'time': time_val.isoformat(),
            'Texhaust': apply_calibration(row_data[1], cal['t1_m'], cal['t1_c']),
            'RHexhaust': apply_calibration(row_data[2], cal['h1_m'], cal['h1_c']),
            'Pressure': row_data[3] or 0.0,
            'Imotor': imotor,
            'type': dev_type,
            'calibrated': True,
            'status': operational_status,
            'running_status': running_status,
            'data_stale': is_stale,
            'is_offline': is_offline,
            'has_data': True,
            'baseline_configured': baseline_configured
        })
    else:
        icomp = row_data[6] or 0
        running_status = 'idle' if is_stale else ('running' if icomp >= 0.4 else 'idle')
        t1c = apply_calibration(row_data[1], cal['t1_m'], cal['t1_c'])
        t2c = apply_calibration(row_data[3], cal['t2_m'], cal['t2_c'])
        h1c = apply_calibration(row_data[2], cal['h1_m'], cal['h1_c'])
        h2c = apply_calibration(row_data[4], cal['h2_m'], cal['h2_c'])
        return jsonify({
            'time': time_val.isoformat(),
            'Treturn': t1c,
            'RHreturn': h1c,
            'Tsupply': t2c,
            'RHsupply': h2c,
            'Tcoil': apply_calibration(row_data[5], cal['tcoil_m'], cal['tcoil_c']),
            'Icompressor': icomp,
            'DeltaT': round(abs(t1c - t2c), 2),
            'DeltaRH': round(abs(h1c - h2c), 2),
            'type': dev_type,
            'calibrated': is_calibrated,
            'status': operational_status,
            'running_status': running_status,
            'data_stale': is_stale,
            'is_offline': is_offline,
            'has_data': True,
            'baseline_configured': baseline_configured
        })

@app.route('/api/device/<int:appliance_id>/latest_n')
@login_required
def api_device_latest_n(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("SELECT operational_status, type FROM appliances WHERE id = %s", (appliance_id,))
    status_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not status_row or ("Dryer" not in status_row[1] and status_row[0] in ['calibration_needed', 'calibrating']):
        return jsonify([]), 200
    limit = int(request.args.get('limit', 1080))
    start = request.args.get('start')
    end = request.args.get('end')
    rows, dev_type = latest_n_rows_for_appliance(appliance_id, limit, start, end)
    cal = get_appliance_calibration(appliance_id)
    if 't1_m' not in cal: return jsonify([])
    result = []
    for r in rows:
        time_val = r[0]
        if "Dryer" in dev_type:
            result.append({
                'time': time_val.isoformat(),
                'Texhaust': apply_calibration(r[1], cal['t1_m'], cal['t1_c']),
                'RHexhaust': apply_calibration(r[2], cal['h1_m'], cal['h1_c']),
                'Pressure': r[3] or 0.0,
                'Imotor': max(0.0, r[4] or 0)
            })
        else:
            t1c = apply_calibration(r[1], cal['t1_m'], cal['t1_c'])
            t2c = apply_calibration(r[3], cal['t2_m'], cal['t2_c'])
            h1c = apply_calibration(r[2], cal['h1_m'], cal['h1_c'])
            h2c = apply_calibration(r[4], cal['h2_m'], cal['h2_c'])
            result.append({
                'time': time_val.isoformat(),
                'Treturn': t1c,
                'RHreturn': h1c,
                'Tsupply': t2c,
                'RHsupply': h2c,
                'Tcoil': apply_calibration(r[5], cal['tcoil_m'], cal['tcoil_c']),
                'Icompressor': r[6] or 0,
                'DeltaT': round(abs(t1c - t2c), 2),
                'DeltaRH': round(abs(h1c - h2c), 2)
            })
    return jsonify(result)

@app.route('/api/device/<int:appliance_id>/table_data')
@login_required
def get_table_data(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("SELECT operational_status, type FROM appliances WHERE id = %s", (appliance_id,))
    status_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not status_row or ("Dryer" not in status_row[1] and status_row[0] in ['calibration_needed', 'calibrating']):
        return jsonify([]), 200
    rows, dev_type = latest_n_rows_for_appliance(appliance_id, 120)
    cal = get_appliance_calibration(appliance_id)
    if 't1_m' not in cal: return jsonify([])
    result = []
    for r in reversed(rows):
        time_val = r[0]
        t1 = apply_calibration(r[1], cal['t1_m'], cal['t1_c'])
        h1 = apply_calibration(r[2], cal['h1_m'], cal['h1_c'])
        if "Dryer" in dev_type:
            result.append({
                'time': time_val.isoformat(),
                'Texhaust': round(t1, 2),
                'RHexhaust': round(h1, 2),
                'Pressure': round(r[3] or 0.0, 2),
                'Imotor': round(max(0.0, r[4] or 0), 2)
            })
        else:
            t2 = apply_calibration(r[3], cal['t2_m'], cal['t2_c'])
            h2 = apply_calibration(r[4], cal['h2_m'], cal['h2_c'])
            t3 = apply_calibration(r[5], cal['tcoil_m'], cal['tcoil_c'])
            amp = r[6] or 0
            result.append({
                'time': time_val.isoformat(),
                'Treturn': round(t1, 2),
                'Tsupply': round(t2, 2),
                'Tcoil': round(t3, 2),
                'RHreturn': round(h1, 2),
                'RHsupply': round(h2, 2),
                'Icompressor': round(amp, 2),
                'DeltaT': round(abs(t1 - t2), 2),
                'DeltaRH': round(abs(h1 - h2), 2)
            })
    return jsonify(result)

@app.route('/api/device/<int:appliance_id>/maintenance_logs')
@login_required
def get_maintenance_logs(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s LIMIT 1", (appliance_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); release_conn(conn)
        return jsonify([]), 200
    cur.execute("""
        SELECT DISTINCT se.timestamp
        FROM sensor_events se
        JOIN sensor_nodes sn ON se.sensor_node_mac = sn.mac_address
        JOIN appliances a ON sn.appliance_id = a.id
        WHERE se.sensor_node_mac = %s
          AND se.event_type = 'maintenance'
          AND se.timestamp >= a.created_at
        ORDER BY se.timestamp DESC
    """, (row[0],))
    rows = cur.fetchall()
    cur.close(); release_conn(conn)
    return jsonify([r[0].isoformat() for r in rows])

# --- API: EXPORT EXCEL ---
@app.route('/api/device/<int:appliance_id>/export_excel')
@login_required
def export_excel(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db error'}), 500
    cur = conn.cursor()
    cur.execute("SELECT name, type, sub_type, operational_status, created_at FROM appliances WHERE id = %s", (appliance_id,))
    app_info = cur.fetchone()
    if not app_info:
        cur.close(); release_conn(conn)
        return jsonify({'error': 'Not found'}), 404
    app_name, dev_type, sub_type, status, created_at = app_info
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    query_params = [appliance_id]
    date_filter = ""
    if start_date:
        date_filter += " AND r.time >= %s"
        query_params.append(start_date)
    if end_date:
        date_filter += " AND r.time <= %s"
        query_params.append(end_date)
    if "Dryer" in dev_type:
        query = f"""
            SELECT r.time, r.texhaust, r.rh_exhaust, r.pressure, r.imotor
            FROM dryer_readings r
            JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND r.time >= a.created_at {date_filter}
            ORDER BY r.time ASC
        """
    else:
        query = f"""
            SELECT r.time, r.treturn, r.rhreturn, r.tsupply, r.rhsupply, r.tcoil, r.icompressor
            FROM hvac_readings r
            JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
            JOIN appliances a ON a.id = sn.appliance_id
            WHERE sn.appliance_id = %s AND r.time >= a.created_at {date_filter}
            ORDER BY r.time ASC
        """
    cur.execute(query, tuple(query_params))
    readings = cur.fetchall()
    cal = get_appliance_calibration(appliance_id)
    cur.close(); release_conn(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sensor Data"
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Device: {app_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A2:D2')
    if "Dryer" in dev_type:
        ws['A2'] = f"Type: {dev_type}"
    else:
        ws['A2'] = f"Type: {dev_type} ({sub_type if sub_type else ''})"
    ws.merge_cells('A3:D3')
    ws['A3'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A4:D4')
    ws['A4'] = f"Data Points: {len(readings)}"
    row_idx = 6
    if "Dryer" in dev_type:
        headers = ["Timestamp", "Exhaust Temp (°C)", "Exhaust RH (%)", "Pressure (hPa)", "Current (A)"]
    else:
        headers = ["Timestamp", "Return Temp (°C)", "Supply Temp (°C)", "Coil Temp (°C)", "Return RH (%)", "Supply RH (%)", "Current (A)", "Delta-T (°C)", "Delta-RH (%)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    if not readings:
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(headers))
        no_data_cell = ws.cell(row=7, column=1, value="No sensor data available for this device in the selected range.")
        no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_data_cell.font = Font(italic=True, color="999999")
    else:
        for r in readings:
            row_idx += 1
            t1 = apply_calibration(r[1], cal['t1_m'], cal['t1_c'])
            h1 = apply_calibration(r[2], cal['h1_m'], cal['h1_c'])
            ws.cell(row=row_idx, column=1, value=r[0].strftime('%Y-%m-%d %H:%M:%S'))
            if "Dryer" in dev_type:
                ws.cell(row=row_idx, column=2, value=round(t1, 2))
                ws.cell(row=row_idx, column=3, value=round(h1, 2))
                ws.cell(row=row_idx, column=4, value=round(r[3] or 0.0, 2))
                ws.cell(row=row_idx, column=5, value=round(r[4] or 0, 2))
            else:
                t2 = apply_calibration(r[3], cal['t2_m'], cal['t2_c'])
                h2 = apply_calibration(r[4], cal['h2_m'], cal['h2_c'])
                t3 = apply_calibration(r[5], cal['tcoil_m'], cal['tcoil_c'])
                amp = r[6] or 0
                ws.cell(row=row_idx, column=2, value=round(t1, 2))
                ws.cell(row=row_idx, column=3, value=round(t2, 2))
                ws.cell(row=row_idx, column=4, value=round(t3, 2))
                ws.cell(row=row_idx, column=5, value=round(h1, 2))
                ws.cell(row=row_idx, column=6, value=round(h2, 2))
                ws.cell(row=row_idx, column=7, value=round(amp, 2))
                ws.cell(row=row_idx, column=8, value=round(abs(t1 - t2), 2))
                ws.cell(row=row_idx, column=9, value=round(abs(h1 - h2), 2))
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx2 in range(6, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx2, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    conn = get_conn()
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT mac_address FROM sensor_nodes WHERE appliance_id = %s LIMIT 1", (appliance_id,))
        mac_row = cur.fetchone()
        maintenance_rows = []
        if mac_row:
            cur.execute("""
                SELECT timestamp FROM sensor_events
                WHERE sensor_node_mac = %s AND event_type = 'maintenance'
                  AND timestamp >= (SELECT created_at FROM appliances WHERE id = %s)
                ORDER BY timestamp DESC
            """, (mac_row[0], appliance_id))
            maintenance_rows = cur.fetchall()
        cur.close()
        release_conn(conn)
        ws2 = wb.create_sheet(title="Maintenance Log")
        ws2['A1'] = f"Device: {app_name}"
        ws2['A1'].font = Font(size=14, bold=True)
        ws2['A3'] = "Maintenance Date"
        ws2['A3'].fill = header_fill
        ws2['A3'].font = header_font
        ws2['A3'].alignment = Alignment(horizontal="center")
        row_idx2 = 3
        for r_maint in maintenance_rows:
            row_idx2 += 1
            ws2.cell(row=row_idx2, column=1, value=r_maint[0].strftime('%Y-%m-%d %H:%M:%S'))
        ws2.column_dimensions['A'].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{app_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# --- API: CALIBRATION PROGRESS ---
@app.route('/api/device/<int:appliance_id>/calibration_progress')
@login_required
def api_calibration_progress(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT operational_status, type, created_at FROM appliances WHERE id = %s", (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not row: return jsonify({'error': 'not found'}), 404
    status, dev_type, created_at = row[0], row[1], row[2]
    if status != 'calibrating':
        return jsonify({'error': 'Not calibrating'}), 400
    tracker = CALIBRATION_TRACKER.get(appliance_id, {})
    start_tcoil = tracker.get('start_tcoil', 25.0)
    conn = get_conn()
    if not conn: return jsonify({'start_tcoil': start_tcoil, 'current_tcoil': None}), 200
    cur = conn.cursor()
    cur.execute("""
        SELECT sr.tcoil FROM hvac_readings sr
        JOIN sensor_nodes sn ON sr.sensor_node_id = sn.id
        JOIN appliances a ON a.id = sn.appliance_id
        WHERE sn.appliance_id = %s AND sr.time >= a.created_at
        ORDER BY sr.time DESC LIMIT 1
    """, (appliance_id,))
    tcoil_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    current_tcoil = float(tcoil_row[0]) if tcoil_row and tcoil_row[0] is not None else start_tcoil
    drop = start_tcoil - current_tcoil
    return jsonify({
        'start_tcoil': round(start_tcoil, 2),
        'current_tcoil': round(current_tcoil, 2),
        'drop': round(drop, 2),
        'target': 8.0,
        'progress_pct': min(100, max(0, round((drop / 8.0) * 100, 1)))
    })

# --- API: SPC LIMITS (reads from spc_manual_baselines) ---
@app.route('/api/device/<int:appliance_id>/spc_limits')
@login_required
def api_spc_limits(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT type, sub_type, operational_status, baseline_configured FROM appliances WHERE id = %s", (appliance_id,))
    db_row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not db_row: return jsonify({'error': 'not found'}), 404
    dev_type, subtype, status, baseline_configured = db_row[0], db_row[1], db_row[2], db_row[3]
    baselines = get_spc_baselines(appliance_id)
    is_inverter = (subtype == 'inverter')

    def make_limit(metric_name):
        if metric_name in baselines:
            b = baselines[metric_name]
            return {"mean": b['mean'], "ucl": b['ucl'], "lcl": b['lcl']}
        return {"mean": 0, "ucl": 0, "lcl": 0}

    if "Dryer" in dev_type:
        return jsonify({
            "type": dev_type,
            "subtype": subtype,
            "status": status,
            "baseline_configured": baseline_configured,
            "temp": make_limit('texhaust'),
            "humidity": make_limit('rhexhaust'),
            "pressure": make_limit('pressure'),
            "current": make_limit('current')
        })
    else:
        return jsonify({
            "type": dev_type,
            "subtype": subtype,
            "status": status,
            "baseline_configured": baseline_configured,
            "deltat": make_limit('deltat'),
            "deltarh": make_limit('deltarh'),
            "tcoil": make_limit('tcoil'),
            "rhreturn": make_limit('rhreturn'),
            "rhsupply": make_limit('rhsupply'),
            "current": make_limit('current')
        })

# --- API: BASELINE ANALYSIS (reads from spc_manual_baselines) ---
@app.route('/api/device/<int:appliance_id>/baseline_analysis')
@login_required
def api_baseline_analysis(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT type, sub_type, operational_status, baseline_configured FROM appliances WHERE id = %s", (appliance_id,))
    db_row = cur.fetchone()
    # Get most recent baseline update timestamp
    cur.execute("SELECT MAX(updated_at) FROM spc_manual_baselines WHERE appliance_id = %s", (appliance_id,))
    ts_row = cur.fetchone()
    baseline_set_at = ts_row[0].isoformat() if ts_row and ts_row[0] else None
    cur.close()
    release_conn(conn)
    if not db_row: return jsonify({'error': 'not found'}), 404
    dev_type, subtype, status, baseline_configured = db_row[0], db_row[1], db_row[2], db_row[3]
    baselines = get_spc_baselines(appliance_id)

    def metric_info(name):
        if name in baselines:
            b = baselines[name]
            return {"mean": b['mean'], "ucl": b['ucl'], "lcl": b['lcl']}
        return None

    if "Dryer" in dev_type:
        return jsonify({
            "type": dev_type,
            "status": status,
            "baseline_configured": baseline_configured,
            "baseline_set_at": baseline_set_at,
            "texhaust": metric_info('texhaust'),
            "rhexhaust": metric_info('rhexhaust'),
            "pressure": metric_info('pressure'),
            "current": metric_info('current')
        })
    else:
        return jsonify({
            "type": dev_type,
            "status": status,
            "baseline_configured": baseline_configured,
            "baseline_set_at": baseline_set_at,
            "deltat": metric_info('deltat'),
            "deltarh": metric_info('deltarh'),
            "tcoil": metric_info('tcoil'),
            "rhreturn": metric_info('rhreturn'),
            "rhsupply": metric_info('rhsupply'),
            "current": metric_info('current')
        })

# --- API: BASELINE CONFIG (NEW) ---
@app.route('/api/device/<int:appliance_id>/baseline_config', methods=['GET', 'POST'])
@login_required
def api_baseline_config(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    cur = conn.cursor()
    cur.execute("SELECT type, operational_status FROM appliances WHERE id = %s", (appliance_id,))
    row = cur.fetchone()
    cur.close()
    release_conn(conn)
    if not row: return jsonify({'error': 'not found'}), 404
    dev_type, status = row[0], row[1]

    if request.method == 'GET':
        baselines = get_spc_baselines(appliance_id)
        expected = DRYER_METRICS if "Dryer" in dev_type else HVAC_METRICS
        result = {}
        for m in expected:
            if m in baselines:
                result[m] = baselines[m]
            else:
                result[m] = {'ucl': '', 'lcl': '', 'mean': ''}
        return jsonify({'type': dev_type, 'metrics': result})

    # POST
    data = request.get_json() or {}
    metrics = data.get('metrics', {})
    expected = DRYER_METRICS if "Dryer" in dev_type else HVAC_METRICS
    baselines_to_save = {}
    for m in expected:
        if m in metrics:
            try:
                ucl = float(metrics[m].get('ucl'))
                lcl = float(metrics[m].get('lcl'))
                if ucl <= lcl:
                    return jsonify({'error': f'UCL must be greater than LCL for {m}'}), 400
                baselines_to_save[m] = {'ucl': ucl, 'lcl': lcl}
            except (ValueError, TypeError):
                return jsonify({'error': f'Invalid values for {m}'}), 400
    if not baselines_to_save:
        return jsonify({'error': 'No valid baseline data provided'}), 400
    success, msg = save_spc_baselines(appliance_id, baselines_to_save)
    if success:
        notify_node_baseline_set(appliance_id)
        return jsonify({'success': True, 'message': msg})
    return jsonify({'error': msg}), 500

# --- API: THRESHOLDS ---
@app.route('/api/device/<int:appliance_id>/thresholds', methods=['GET', 'POST'])
@login_required
def api_thresholds(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    if request.method == 'GET':
        cur = conn.cursor()
        cur.execute("SELECT alert_rhexhaust_threshold, alert_enabled FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        cur.close()
        release_conn(conn)
        if not row: return jsonify({'error': 'not found'}), 404
        return jsonify({
            "alert_rhexhaust_threshold": float(row[0]) if row[0] else 40.0,
            "alert_enabled": row[1] if row[1] is not None else True
        })
    try:
        cur = conn.cursor()
        data = request.get_json() or {}
        threshold = data.get('alert_rhexhaust_threshold')
        enabled = data.get('alert_enabled')
        updates = []
        params = []
        if threshold is not None:
            updates.append("alert_rhexhaust_threshold = %s")
            params.append(float(threshold))
        if enabled is not None:
            updates.append("alert_enabled = %s")
            params.append(bool(enabled))
        if updates:
            params.append(appliance_id)
            cur.execute(f"UPDATE appliances SET {', '.join(updates)} WHERE id = %s", params)
            conn.commit()
        cur.close()
        release_conn(conn)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- API: ALERTS ---
@app.route('/api/device/<int:appliance_id>/alerts')
@login_required
def api_alerts(appliance_id):
    conn = get_conn()
    if not conn: return jsonify([]), 500
    cur = conn.cursor()
    cur.execute("""
        SELECT a.id, a.alert_type, a.message, a.value, a.threshold, a.created_at, a.resolved_at, a.acknowledged
        FROM alerts a
        WHERE a.appliance_id = %s
        ORDER BY a.created_at DESC LIMIT 50
    """, (appliance_id,))
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    result = []
    for r in rows:
        result.append({
            "id": r[0], "alert_type": r[1], "message": r[2],
            "value": float(r[3]) if r[3] is not None else None,
            "threshold": float(r[4]) if r[4] is not None else None,
            "created_at": r[5].isoformat() if r[5] else None,
            "resolved_at": r[6].isoformat() if r[6] else None,
            "acknowledged": r[7]
        })
    return jsonify(result)

# --- API: HVAC ANALYTICS ---
@app.route('/api/device/<int:appliance_id>/hvac_analytics')
@login_required
def hvac_analytics(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    try:
        cur = conn.cursor()
        cur.execute("SELECT type FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        if not row or "HVAC" not in row[0]:
            return jsonify({'error': 'Not HVAC'}), 400
        start = request.args.get('start')
        end = request.args.get('end')
        if start and end:
            cur.execute("""
                SELECT DATE(r.time) as date,
                       AVG(r.treturn), AVG(r.tsupply), AVG(r.tcoil)
                FROM hvac_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.icompressor >= 0.4 AND r.time >= a.created_at AND r.time >= %s AND r.time <= %s
                GROUP BY DATE(r.time)
                ORDER BY DATE(r.time) DESC
                LIMIT 30
            """, (appliance_id, start, end))
        else:
            cur.execute("""
                SELECT DATE(r.time) as date,
                       AVG(r.treturn), AVG(r.tsupply), AVG(r.tcoil)
                FROM hvac_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.icompressor >= 0.4 AND r.time >= a.created_at
                GROUP BY DATE(r.time)
                ORDER BY DATE(r.time) DESC
                LIMIT 30
            """, (appliance_id,))
        readings = cur.fetchall()
        result = []
        for r in readings:
            result.append({
                "date": r[0].isoformat(),
                "avg_intake": round(r[1], 2),
                "avg_exit": round(r[2], 2),
                "avg_coil": round(r[3], 2)
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cur.close()
            release_conn(conn)

# --- API: DRYER ANALYTICS ---
@app.route('/api/device/<int:appliance_id>/dryer_analytics')
@login_required
def dryer_analytics(appliance_id):
    conn = get_conn()
    if not conn: return jsonify({'error': 'db'}), 500
    try:
        cur = conn.cursor()
        cur.execute("SELECT type FROM appliances WHERE id = %s", (appliance_id,))
        row = cur.fetchone()
        if not row or "Dryer" not in row[0]:
            return jsonify({'error': 'Not a dryer'}), 400
        cycle_start = 0.4
        cycle_end = 0.15
        prominence_threshold = 0.5
        start = request.args.get('start')
        end = request.args.get('end')
        if start and end:
            cur.execute("""
                SELECT r.time, r.texhaust, r.rh_exhaust, r.pressure, r.imotor
                FROM dryer_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.time >= a.created_at AND r.time >= %s AND r.time <= %s
                ORDER BY r.time ASC
            """, (appliance_id, start, end))
        else:
            cur.execute("""
                SELECT r.time, r.texhaust, r.rh_exhaust, r.pressure, r.imotor
                FROM dryer_readings r
                JOIN sensor_nodes sn ON r.sensor_node_id = sn.id
                JOIN appliances a ON a.id = sn.appliance_id
                WHERE sn.appliance_id = %s AND r.time >= a.created_at
                ORDER BY r.time ASC
            """, (appliance_id,))
        readings = cur.fetchall()
        if not readings:
            return jsonify([])
        cycles = []
        in_cycle = False
        current_cycle = {}
        _prev_current = 0.0
        _peak_state = "IDLE"
        _peak_max = 0.0
        _peak_valley = 0.0
        _peak_values = []

        def _confirm_peak():
            nonlocal _peak_state, _peak_max, _peak_valley
            if _peak_state in ("RISING", "FALLING") and _peak_max > 0:
                prominence = _peak_max - _peak_valley
                if prominence >= prominence_threshold:
                    _peak_values.append(_peak_max)
            _peak_state = "IDLE"
            _peak_max = 0.0
            _peak_valley = 0.0

        for i, r in enumerate(readings):
            time_val, tex, rhex, press, imotor = r
            imotor = float(imotor) if imotor is not None else 0.0
            tex = float(tex) if tex is not None else 0.0
            rhex = float(rhex) if rhex is not None else None
            if in_cycle and i > 0:
                gap = (time_val - readings[i-1][0]).total_seconds()
                if gap > 60:
                    in_cycle = False
                    prev_time = readings[i-1][0]
                    current_cycle["end_time"] = prev_time
                    duration = current_cycle["end_time"] - current_cycle["start_time"]
                    current_cycle["duration_minutes"] = round(duration.total_seconds() / 60, 1)
                    valid_rh = [v for v in current_cycle["_rh_history"] if v is not None]
                    last_10 = valid_rh[-10:] if len(valid_rh) > 10 else valid_rh
                    current_cycle["end_rh_avg"] = round(sum(last_10) / len(last_10), 2) if last_10 else current_cycle.get("start_rh", 0)
                    currents = current_cycle["_currents"]
                    current_cycle["current_consumption"] = round(sum(currents), 2)
                    current_cycle["current_spike_avg"] = round(sum(_peak_values) / len(_peak_values), 2) if _peak_values else 0.0
                    current_cycle["ignition_count"] = len(_peak_values)
                    del current_cycle["_rh_history"]
                    del current_cycle["_currents"]
                    current_cycle["start_time"] = current_cycle["start_time"].isoformat()
                    current_cycle["end_time"] = current_cycle["end_time"].isoformat()
                    cycles.append(current_cycle)
                    current_cycle = {}
                    _confirm_peak()
                    _prev_current = 0.0
                    _peak_values = []
            if imotor > cycle_start and not in_cycle:
                in_cycle = True
                current_cycle = {
                    "start_time": time_val,
                    "min_temp": tex,
                    "max_temp": tex,
                    "start_rh": rhex,
                    "_rh_history": [],
                    "_currents": []
                }
                _prev_current = imotor
                _confirm_peak()
                _peak_values = []
            if in_cycle:
                current_cycle["max_temp"] = max(current_cycle["max_temp"], tex)
                current_cycle["min_temp"] = min(current_cycle["min_temp"], tex)
                if rhex is not None:
                    current_cycle["_rh_history"].append(rhex)
                current_cycle["_currents"].append(imotor)
                if imotor > _prev_current:
                    if _peak_state in ("IDLE", "FALLING"):
                        if _peak_state == "FALLING":
                            _confirm_peak()
                        _peak_valley = _prev_current
                    _peak_state = "RISING"
                    if imotor > _peak_max:
                        _peak_max = imotor
                elif imotor < _prev_current:
                    if _peak_state == "RISING":
                        _peak_state = "FALLING"
                _prev_current = imotor
            if imotor < cycle_end and in_cycle:
                in_cycle = False
                current_cycle["end_time"] = time_val
                duration = current_cycle["end_time"] - current_cycle["start_time"]
                current_cycle["duration_minutes"] = round(duration.total_seconds() / 60, 1)
                valid_rh = [v for v in current_cycle["_rh_history"] if v is not None]
                last_10 = valid_rh[-10:] if len(valid_rh) > 10 else valid_rh
                current_cycle["end_rh_avg"] = round(sum(last_10) / len(last_10), 2) if last_10 else current_cycle.get("start_rh", 0)
                currents = current_cycle["_currents"]
                current_cycle["current_consumption"] = round(sum(currents), 2)
                current_cycle["current_spike_avg"] = round(sum(_peak_values) / len(_peak_values), 2) if _peak_values else 0.0
                current_cycle["ignition_count"] = len(_peak_values)
                del current_cycle["_rh_history"]
                del current_cycle["_currents"]
                current_cycle["start_time"] = current_cycle["start_time"].isoformat()
                current_cycle["end_time"] = current_cycle["end_time"].isoformat()
                cycles.append(current_cycle)
                current_cycle = {}
                _confirm_peak()
                _prev_current = 0.0
                _peak_values = []
        if in_cycle:
            in_cycle = False
            last_r = readings[-1]
            current_cycle["end_time"] = last_r[0]
            duration = current_cycle["end_time"] - current_cycle["start_time"]
            current_cycle["duration_minutes"] = round(duration.total_seconds() / 60, 1)
            valid_rh = [v for v in current_cycle["_rh_history"] if v is not None]
            last_10 = valid_rh[-10:] if len(valid_rh) > 10 else valid_rh
            current_cycle["end_rh_avg"] = round(sum(last_10) / len(last_10), 2) if last_10 else current_cycle.get("start_rh", 0)
            _confirm_peak()
            currents = current_cycle["_currents"]
            current_cycle["current_consumption"] = round(sum(currents), 2)
            current_cycle["current_spike_avg"] = round(sum(_peak_values) / len(_peak_values), 2) if _peak_values else 0.0
            current_cycle["ignition_count"] = len(_peak_values)
            del current_cycle["_rh_history"]
            del current_cycle["_currents"]
            current_cycle["start_time"] = current_cycle["start_time"].isoformat()
            current_cycle["end_time"] = current_cycle["end_time"].isoformat()
            cycles.append(current_cycle)
            current_cycle = {}
        cycles = [c for c in cycles if c.get("duration_minutes", 0) >= 1.0]
        return jsonify(cycles)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cur.close()
            release_conn(conn)

# --- MAIN ---
if __name__ == '__main__':
    mqtt_client.connect(MQTT_HOST, MQTT_PORT, 60)
    mqtt_client.loop_start()
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
